#!/usr/bin/env python3
"""YSL VIVI Monitoring xlsx 빌더.

Kiehl's VIVI Monitoring 양식을 미러링하면서 YSL 데이터로 4시트 생성:
  00. Keywords&MSV  — 1:1 검색 키워드 360개 + Google/Naver MSV
  01. Query List    — 360 질문 + 카테고리/인텐트/축1·2·3
  02. Mention       — 4 사이클 통합 응답 + 12 브랜드 멘션
  03. Citation      — 4 사이클 통합 인용 + YSL 자사 도메인 플래그

사용:
  python3 ysl/scripts/build_vivi_monitoring.py
"""
import datetime
import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# =============================================================================
# 경로
# =============================================================================
ROOT = Path(__file__).resolve().parents[1]
QLIST = ROOT / 'final' / 'Bubbleshare_YSL_Question_List_수정.xlsx'
BRANDS = ROOT / 'template' / 'brands_config.json'
KW_DIR = ROOT / 'ysl_질문 수집' / 'keyword'
Q_DIR = ROOT / 'ysl_질문 수집' / 'question'
CITATION_ANALYSIS_DIR = ROOT / 'ysl_질문 수집' / 'citation_analysis'
OUT_XLSX = ROOT / 'final' / 'Bubbleshare_YSL_VIVI_Monitoring.xlsx'

# BackOffice가 citation_analysis CSV로 채워주는 분류·메타 컬럼들
CLASSIFY_COLS = [
    'Domain Type', 'domain CI', 'Content Type', 'content CI',
    'Published Date', 'Date CI', 'Publisher Name', 'Name CI',
    'Meta Author', 'Meta Site Name', 'Meta Language', 'Meta Description',
    'Meta OG Type', 'Meta Keywords',
]

PREFIX_MAP = {'향수': 'PF', '기프팅': 'GF', '쿠션': 'CS'}

HEADER_FILL = PatternFill('solid', fgColor='7030A0')
HEADER_FONT = Font(bold=True, color='FFFFFF')
SUBHEADER_FILL = PatternFill('solid', fgColor='ECD1ED')


# =============================================================================
# 1. 메타 로딩
# =============================================================================

def load_brands():
    cfg = json.loads(BRANDS.read_text(encoding='utf-8'))
    ysl_name = cfg['brand']['name']
    ysl_url = cfg['brand'].get('url', 'yslbeauty')
    brands = {ysl_name: cfg['brand']['variations']}
    for c in cfg['competitors']:
        brands[c['name']] = c['variations']
    return brands, ysl_name, ysl_url


def _s(v):
    if isinstance(v, str):
        return v.strip()
    return ''


def load_meta():
    """`Questions (Final) + KW`(메타·KW·MSV)와 `Questions (Final)`(축1/2/3) 시트 join."""
    wb = load_workbook(QLIST, data_only=True)
    ws_kw = wb['Questions (Final) + KW']
    ws_f = wb['Questions (Final)']

    final_meta = {}  # No -> axis1/2/3 + verify
    for r in range(5, ws_f.max_row + 1):
        no = ws_f.cell(r, 2).value
        if not isinstance(no, (int, float)):
            continue
        n = int(no)
        if not (1 <= n <= 360):
            continue
        final_meta[n] = {
            'verify': _s(ws_f.cell(r, 7).value),
            'axis1': _s(ws_f.cell(r, 8).value),
            'axis2': _s(ws_f.cell(r, 9).value),
            'axis3': _s(ws_f.cell(r, 10).value),
        }

    rows = []
    for r in range(5, ws_kw.max_row + 1):
        no = ws_kw.cell(r, 2).value
        if not isinstance(no, (int, float)):
            continue
        n = int(no)
        cat = _s(ws_kw.cell(r, 3).value)
        gmsv = ws_kw.cell(r, 10).value or 0
        nmsv = ws_kw.cell(r, 11).value or 0
        ax = final_meta.get(n, {})
        prefix = PREFIX_MAP.get(cat, 'XX')
        rows.append({
            'no': n,
            'category': cat,
            'sub': _s(ws_kw.cell(r, 4).value),
            'intent': _s(ws_kw.cell(r, 5).value),
            'question': _s(ws_kw.cell(r, 6).value),
            'keyword': _s(ws_kw.cell(r, 7).value),
            'gmsv': int(gmsv) if isinstance(gmsv, (int, float)) else 0,
            'nmsv': int(nmsv) if isinstance(nmsv, (int, float)) else 0,
            'axis1': ax.get('axis1', ''),
            'axis2': ax.get('axis2', ''),
            'axis3': ax.get('axis3', ''),
            'verify': ax.get('verify', ''),
            'ref_q': f'{prefix}-{n:03d}',
            'ref_kw': f'{prefix}-KW-{n:03d}',
        })
    return rows


# =============================================================================
# 2. CSV 로딩
# =============================================================================

def parse_epoch_from_filename(fname):
    m = re.search(r'-(\d{13})\.csv$', fname)
    return int(m.group(1)) if m else None


def epoch_to_dt(ms):
    if ms is None:
        return None
    return datetime.datetime.fromtimestamp(ms / 1000)


def _load_csvs(files, source_kind):
    dfs = []
    for i, f in enumerate(sorted(files), 1):
        df = pd.read_csv(f, encoding='utf-8-sig')
        df.columns = [c.strip().lstrip('﻿') for c in df.columns]
        ms = parse_epoch_from_filename(f.name)
        df['_cycle'] = i
        df['_date'] = epoch_to_dt(ms)
        df['_source_kind'] = source_kind
        df['_file'] = f.name
        dfs.append(df)
    return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()


def load_responses_long():
    kw = _load_csvs(KW_DIR.glob('ysl_kw_ai_overview-*.csv'), 'keyword')
    q = _load_csvs(Q_DIR.glob('ysl_questions_ai_overview-*.csv'), 'question')
    return pd.concat([kw, q], ignore_index=True)


def load_citations_long():
    kw = _load_csvs(KW_DIR.glob('*ysl_kw_citations*.csv'), 'keyword')
    q = _load_csvs(Q_DIR.glob('*ysl_questions_citations*.csv'), 'question')
    return pd.concat([kw, q], ignore_index=True)


def load_citation_classification():
    """citation_analysis/*.csv → {(reference_id, url): {col: val, ...}} 매핑.

    BackOffice가 분류해서 보낸 인용 메타(Domain Type, Content Type 등)를
    (Reference ID, URL) 단위로 lookup 가능하게 한다. 폴더 없거나 비었으면
    빈 dict 반환 → 기존 NaN 동작 유지 (backwards compatible).
    """
    if not CITATION_ANALYSIS_DIR.exists():
        return {}

    lookup = {}
    files = sorted(CITATION_ANALYSIS_DIR.glob('*.csv'))
    for f in files:
        try:
            df = pd.read_csv(f, encoding='utf-8-sig')
            df.columns = [c.strip().lstrip('﻿') for c in df.columns]
        except Exception as e:
            print(f'  citation_analysis 파일 로딩 실패 ({f.name}): {e}')
            continue
        for _, row in df.iterrows():
            ref = str(row.get('reference_id', '') or '').strip()
            url = str(row.get('URL', '') or '').strip()
            if not ref or not url:
                continue
            entry = {}
            for col in CLASSIFY_COLS:
                if col not in df.columns:
                    continue
                v = row.get(col)
                if v is None or (isinstance(v, float) and pd.isna(v)):
                    continue
                entry[col] = v
            if entry:
                lookup[(ref, url)] = entry
    return lookup


# =============================================================================
# 3. 멘션 추출
# =============================================================================

def extract_mentions_with_count(text, brands):
    """각 브랜드 (Y/N, 매칭 횟수). variations는 합산."""
    if not isinstance(text, str) or not text.strip():
        return {b: ('N', 0) for b in brands}
    t = text.lower()
    out = {}
    for brand, variations in brands.items():
        cnt = 0
        for v in variations:
            cnt += t.count(v.lower())
        out[brand] = ('Y' if cnt > 0 else 'N', cnt)
    return out


# =============================================================================
# 4. 시트 빌드
# =============================================================================

def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _autosize(ws, max_w=60):
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 10
        for cell in col[:50]:  # 처음 50행만 보고 결정
            v = cell.value
            if v is None:
                continue
            l = len(str(v))
            if l > max_len:
                max_len = min(l, max_w)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_w)


def build_keywords_sheet(meta_rows, ws):
    """00. Keywords&MSV"""
    headers = ['No.', 'Keyword', 'Google MSV', 'Naver MSV']
    ws.append(headers)
    for r in meta_rows:
        ws.append([r['no'], r['keyword'], r['gmsv'], r['nmsv']])
    _style_header(ws)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def build_query_sheet(meta_rows, ws):
    """01. Query List — Kiehl's 양식 + 축3 신설. 축은 Sub Category 바로 뒤."""
    headers = [
        'No.', 'Reference ID (Q)', 'Reference ID (KW)',
        'Category', 'Sub Category (Seed KW)',
        'Category (Axis 1) - 사용상황/TPO',
        'Category (Axis 2) - 소비자 프로필',
        'Category (Axis 3) - 구매 허들',
        'Intent', 'Question', '1:1 Keyword',
        'Google MSV', 'Naver MSV',
    ]
    ws.append(headers)
    for r in meta_rows:
        ws.append([
            r['no'], r['ref_q'], r['ref_kw'],
            r['category'], r['sub'],
            r['axis1'], r['axis2'], r['axis3'],
            r['intent'], r['question'], r['keyword'],
            r['gmsv'], r['nmsv'],
        ])
    _style_header(ws)
    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = ws.dimensions
    _autosize(ws)


def build_mention_sheet(meta_rows, resp_df, brands, ws):
    """02. Mention — long-format 통합."""
    by_ref = {}
    for r in meta_rows:
        by_ref[r['ref_q']] = (r, 'question')
        by_ref[r['ref_kw']] = (r, 'keyword')

    brand_order = list(brands.keys())  # YSL Beauty 먼저, 경쟁사 11

    headers = [
        'Index', 'Date', 'Cycle', 'Reference ID', 'Keyword/Query',
        'Category', 'Sub Category',
        'Question', '1:1 Keyword',
        'Brand / Non-Brand', 'Google MSV', 'Naver MSV',
        'Channel', 'Has AI Overview', 'AI Response Text',
        'Source URLs', 'Source Titles', 'Source Domains',
        'Original Source URLs (With Highlights)',
    ]
    for b in brand_order:
        headers.append(f'{b} 언급')
        headers.append(f'{b} 언급#')
    ws.append(headers)

    idx = 0
    for _, row in resp_df.iterrows():
        ref_raw = row.get('Reference ID')
        if ref_raw is None or (isinstance(ref_raw, float) and pd.isna(ref_raw)):
            ref_raw = row.get('reference_id', '')
        ref = str(ref_raw).strip() if ref_raw is not None else ''

        meta_lookup = by_ref.get(ref)
        if meta_lookup:
            m, kq = meta_lookup
        else:
            m = {'category': '', 'sub': '', 'intent': '', 'question': '',
                 'keyword': '', 'gmsv': '', 'nmsv': '',
                 'axis1': '', 'axis2': '', 'axis3': ''}
            kq = row.get('_source_kind', '')

        has_ai = str(row.get('Has AI Overview', '')).strip().lower()
        text_raw = row.get('AI Response Text', '')
        text = text_raw if (has_ai in ('yes', 'true') and isinstance(text_raw, str)) else ''
        mentions = extract_mentions_with_count(text, brands)

        idx += 1
        date_v = row.get('_date')
        date_str = date_v.strftime('%Y-%m-%d %H:%M:%S') if isinstance(date_v, datetime.datetime) else ''

        def _cell(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return ''
            return v

        out = [
            idx, date_str, row.get('_cycle', ''), ref, kq,
            m['category'], m['sub'],
            m['question'], m['keyword'],
            'Non-Brand',
            m['gmsv'], m['nmsv'],
            row.get('Channel', ''),
            row.get('Has AI Overview', ''),
            text if isinstance(text, str) else '',
            _cell(row.get('Source URLs', '')),
            _cell(row.get('Source Titles', '')),
            _cell(row.get('Source Domains', '')),
            _cell(row.get('Original Source URLs (With Highlights)', '')),
        ]
        for b in brand_order:
            yn, cnt = mentions[b]
            out.append(yn)
            out.append(cnt)
        ws.append(out)

    _style_header(ws)
    ws.freeze_panes = 'E2'
    ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}1'


def build_citation_sheet(meta_rows, cit_df, ysl_url, ws, classify_lookup=None):
    """03. Citation — 자사 도메인 플래그 + citation_analysis enrichment.

    classify_lookup: (reference_id, url) → {col: val} dict. 매칭되는 행은
    raw CSV 값 대신 classify_lookup 값으로 덮어씀. 매칭 안 되면 raw 사용.
    """
    by_ref = {}
    for r in meta_rows:
        by_ref[r['ref_q']] = r
        by_ref[r['ref_kw']] = r

    classify_lookup = classify_lookup or {}

    headers = [
        'Index', 'Date', 'Cycle', 'Reference ID', 'Keyword/Query',
        'Category', 'Sub Category', 'Intent',
        'Platform', 'Query', 'Rank',
        'Title', 'URL', 'Domain',
        'Domain Type', 'domain CI', 'Content Type', 'content CI',
        'Published Date', 'Date CI', 'Publisher Name', 'Name CI',
        'Meta Author', 'Meta Site Name', 'Meta Language', 'Meta Description',
        'Meta OG Type', 'Meta Keywords',
        'YSL 공홈',
    ]
    ws.append(headers)

    idx = 0
    enriched = 0
    ysl_url_l = ysl_url.lower()

    # openpyxl 금지 control char (0x00-0x08, 0x0B-0x0C, 0x0E-0x1F) 제거용
    _illegal_re = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

    def _cell(v):
        if v is None:
            return ''
        if isinstance(v, float) and pd.isna(v):
            return ''
        if isinstance(v, str):
            return _illegal_re.sub('', v)
        return v

    for _, row in cit_df.iterrows():
        ref = str(row.get('reference_id', '') or '').strip()
        url = str(row.get('URL', '') or '').strip()
        m = by_ref.get(ref, {})
        kq = row.get('_source_kind', '')
        domain_raw = row.get('Domain', '')
        domain_l = str(domain_raw or '').lower()
        ysl_flag = 'Y' if (ysl_url_l in domain_l or 'yslbeauty' in domain_l) else 'N'

        # citation_analysis enrichment: 매칭되면 분류 컬럼들 덮어쓰기
        classify_data = classify_lookup.get((ref, url), {})
        if classify_data:
            enriched += 1

        def _classify(col):
            if col in classify_data:
                return _cell(classify_data[col])
            return _cell(row.get(col, ''))

        idx += 1
        out = [
            idx,
            _cell(row.get('Job Created At', '')),
            row.get('_cycle', ''),
            ref, kq,
            m.get('category', ''), m.get('sub', ''), m.get('intent', ''),
            _cell(row.get('Platform', '')),
            _cell(row.get('Query', '')),
            _cell(row.get('Rank', '')),
            _cell(row.get('Title', '')),
            _cell(row.get('URL', '')),
            _cell(row.get('Domain', '')),
            _classify('Domain Type'),
            _classify('domain CI'),
            _classify('Content Type'),
            _classify('content CI'),
            _classify('Published Date'),
            _classify('Date CI'),
            _classify('Publisher Name'),
            _classify('Name CI'),
            _classify('Meta Author'),
            _classify('Meta Site Name'),
            _classify('Meta Language'),
            _classify('Meta Description'),
            _classify('Meta OG Type'),
            _classify('Meta Keywords'),
            ysl_flag,
        ]
        ws.append(out)

    _style_header(ws)
    ws.freeze_panes = 'E2'
    ws.auto_filter.ref = f'A1:{get_column_letter(ws.max_column)}1'

    if classify_lookup:
        print(f'    citation enrichment: {enriched:,}/{idx:,} ({enriched/idx*100:.1f}%) 행 매칭')


# =============================================================================
# 5. 메인
# =============================================================================

def main():
    print('=== 1. 메타 로딩 ===')
    brands, ysl_name, ysl_url = load_brands()
    print(f'  브랜드: {len(brands)}개 (자사 + {len(brands) - 1} 경쟁사)')

    meta_rows = load_meta()
    print(f'  Question 메타: {len(meta_rows)}개')

    print('\n=== 2. 응답/인용 로딩 ===')
    resp_df = load_responses_long()
    cit_df = load_citations_long()
    print(f'  응답: {len(resp_df)}')
    print(f'  인용: {len(cit_df)}')

    print('\n=== 2.5 citation_analysis 분류 데이터 로딩 ===')
    classify_lookup = load_citation_classification()
    print(f'  분류된 (Reference ID, URL) 조합: {len(classify_lookup):,}')

    print('\n=== 3. xlsx 빌드 ===')
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet('00. Keywords&MSV')
    build_keywords_sheet(meta_rows, ws1)
    print(f'  00. Keywords&MSV: {ws1.max_row - 1} rows')

    ws2 = wb.create_sheet('01. Query List')
    build_query_sheet(meta_rows, ws2)
    print(f'  01. Query List: {ws2.max_row - 1} rows')

    ws3 = wb.create_sheet('02. Mention')
    build_mention_sheet(meta_rows, resp_df, brands, ws3)
    print(f'  02. Mention: {ws3.max_row - 1} rows')

    ws4 = wb.create_sheet('03. Citation')
    build_citation_sheet(meta_rows, cit_df, ysl_url, ws4, classify_lookup)
    print(f'  03. Citation: {ws4.max_row - 1} rows')

    print(f'\n저장: {OUT_XLSX}')
    wb.save(OUT_XLSX)
    print('완료')


if __name__ == '__main__':
    main()
