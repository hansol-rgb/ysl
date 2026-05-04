#!/usr/bin/env python3
"""
Audit Report Data Builder

`Bubbleshare_YSL_VIVI_Monitoring_{suffix}.xlsx` (raw) +
`Bubbleshare_YSL_Question_List_수정.xlsx` (메타) 를 입력받아
PDF 템플릿 (Audit_Report_Template.pdf) 17 페이지의 모든 데이터 sub-block 을
**26 시트 long-format xlsx** 로 추출한다.

설계 근거: projects/ysl/scripts/audit_report_data_plan.md
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# 설정 / 메타
# ---------------------------------------------------------------------------

BRANDS = [
    "YSL Beauty", "Dior", "Chanel", "Hera", "MAC", "Jo Malone",
    "Nars", "Estee Lauder", "Lancome", "Tom Ford", "Sulwhasoo", "Prada Beauty",
]

COMPETITORS = [b for b in BRANDS if b != "YSL Beauty"]

CATEGORIES = ["향수", "기프팅", "쿠션"]
INTENTS = ["니즈 인식", "정보 탐색", "대안 비교", "구매 결정"]
CHANNELS = ["chatgpt", "google", "naver"]
CHANNEL_LABELS = {"chatgpt": "ChatGPT", "google": "Google AIO", "naver": "Naver AI Briefing"}

# 자사 / 경쟁사 자사몰 도메인 키워드 (Domain LIKE %keyword%)
BRAND_OWN_DOMAIN_KEYWORDS: dict[str, list[str]] = {
    "YSL Beauty": ["yslbeauty"],
    "Dior": ["dior.com"],
    "Chanel": ["chanel.com"],
    "Hera": ["hera.com"],
    "MAC": ["maccosmetics"],
    "Jo Malone": ["jomalone"],
    "Nars": ["narscosmetics"],
    "Estee Lauder": ["esteelauder"],
    "Lancome": ["lancome"],
    "Tom Ford": ["tomford"],
    "Sulwhasoo": ["sulwhasoo.com"],
    "Prada Beauty": ["pradabeauty", "prada.com"],
}

# E-commerce 리테일러 화이트리스트 (도메인 키워드 → 라벨)
ECOMMERCE_RETAILERS: list[tuple[str, list[str]]] = [
    ("쿠팡", ["coupang.com"]),
    ("SSG/이마트/신세계몰", ["ssg.com"]),
    ("카카오 선물하기", ["gift.kakao.com"]),
    ("카카오 쇼핑하우", ["shoppinghow.kakao.com"]),
    ("올리브영", ["oliveyoung.co.kr"]),
    ("세포라", ["sephora.com", "sephora.hk"]),
    ("무신사", ["musinsa.com"]),
    ("롯데온", ["lotteon.com"]),
    ("네이버 스마트스토어", ["smartstore.naver.com"]),
    ("11번가", ["11st.co.kr"]),
    ("컬리", ["kurly.com"]),
    ("SSG 면세점", ["ssgdfs.com"]),
    ("롯데 면세점", ["lottedfs.com"]),
    ("G마켓", ["gmarket.co.kr", "cjgmarket.com"]),
]

ECOMMERCE_OTHER_HINTS = ["harrods.com", "idus.com", "kream.co.kr", "fragrantica.com", "ibspot.com"]

# 페이지 유형 URL 패턴
PAGE_TYPE_PATTERNS: list[tuple[str, list[str]]] = [
    ("상품 상세 (PDP)", ["/product/", "/goods/", "/prd/", "/p/", "productid", "goodsid"]),
    ("기획전/매거진", ["/promotion/", "/magazine/", "/event/", "/curation/", "/story/", "/ranking/", "/planshop/"]),
    ("검색결과", ["/search", "/display/", "category", "?query=", "&query=", "?q=", "&q="]),
]

AXIS_LABELS = [
    ("Axis 1", "사용 상황/TPO", "Category (Axis 1) - 사용상황/TPO"),
    ("Axis 2", "소비자 프로필", "Category (Axis 2) - 소비자 프로필"),
    ("Axis 3", "구매 허들", "Category (Axis 3) - 구매 허들"),
]

# ---------------------------------------------------------------------------
# Raw 로딩
# ---------------------------------------------------------------------------

@dataclass
class Raw:
    query_list: list[dict[str, Any]]
    mention: list[dict[str, Any]]
    citation: list[dict[str, Any]]
    kw_msv: list[dict[str, Any]]
    category_purpose: dict[str, str]


def _rows_as_dicts(ws, header_row: int = 1) -> Iterable[dict[str, Any]]:
    headers: list[str] | None = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < header_row:
            continue
        if i == header_row:
            headers = [str(h) if h is not None else f"_c{j}" for j, h in enumerate(row)]
            continue
        yield dict(zip(headers, row))


def load_raw(monitoring_path: Path, question_list_path: Path) -> Raw:
    print(f"[load] {monitoring_path.name}")
    wb = openpyxl.load_workbook(monitoring_path, read_only=True, data_only=True)
    query_list = list(_rows_as_dicts(wb["01. Query List"]))
    mention = list(_rows_as_dicts(wb["02. Mention"]))
    citation = list(_rows_as_dicts(wb["03. Citation"]))
    kw_msv = list(_rows_as_dicts(wb["00. Keywords&MSV"]))
    wb.close()

    print(f"[load] {question_list_path.name} (Prompt setting → Category 목적)")
    category_purpose = _load_category_purpose(question_list_path)
    return Raw(query_list=query_list, mention=mention, citation=citation,
               kw_msv=kw_msv, category_purpose=category_purpose)


def _load_category_purpose(path: Path) -> dict[str, str]:
    """Question_List.xlsx 의 'Prompt setting' 시트 R28-R31 에서 카테고리별 목적 추출."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Prompt setting"]
    purposes: dict[str, str] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # R29~R31 형식: [None, "YSL Beauty"|None, "향수 (40%)", "<목적 텍스트>", 0.4, 144]
        if 28 <= i <= 33:
            cat_cell = row[2] if len(row) > 2 else None
            purpose_cell = row[3] if len(row) > 3 else None
            if cat_cell and purpose_cell:
                cat = str(cat_cell)
                # "향수 (40%)" → "향수"
                m = re.match(r"^(\S+)", cat)
                key = m.group(1) if m else cat
                if key in CATEGORIES:
                    purposes[key] = str(purpose_cell).strip()
    wb.close()
    return purposes


# ---------------------------------------------------------------------------
# Helpers (집계)
# ---------------------------------------------------------------------------

def safe_div(a: float, b: float) -> float:
    return a / b if b else 0.0


def is_yes(val: Any) -> bool:
    return str(val).strip().upper() == "Y" or val == "Yes"


def commercial(row: dict[str, Any]) -> bool:
    """12 브랜드 중 1개라도 언급된 응답"""
    for b in BRANDS:
        if is_yes(row.get(f"{b} 언급")):
            return True
    return False


def has_ai(row: dict[str, Any]) -> bool:
    return str(row.get("Has AI Overview", "")).strip() == "Yes"


def ysl_mention(row: dict[str, Any]) -> bool:
    return is_yes(row.get("YSL Beauty 언급"))


def domain_lower(d: Any) -> str:
    return str(d or "").lower()


def url_lower(u: Any) -> str:
    return str(u or "").lower()


def is_brand_own_domain(domain: str, brand: str) -> bool:
    d = domain.lower()
    return any(kw.lower() in d for kw in BRAND_OWN_DOMAIN_KEYWORDS.get(brand, []))


def ecommerce_retailer(domain: str) -> str | None:
    d = domain.lower()
    for label, kws in ECOMMERCE_RETAILERS:
        if any(kw in d for kw in kws):
            return label
    if any(kw in d for kw in ECOMMERCE_OTHER_HINTS):
        return "기타 이커머스"
    return None


def page_type(url: str) -> str:
    u = url.lower()
    for label, kws in PAGE_TYPE_PATTERNS:
        if any(kw in u for kw in kws):
            return label
    return "기타"


def compute_funnel(rows: list[dict[str, Any]], citation_keys: set[tuple]) -> list[tuple[str, int, float]]:
    """rows (Mention) 에 대해 Funnel 5단계 카운트와 비율."""
    qs = len(rows)
    ai = sum(1 for r in rows if has_ai(r))
    com = sum(1 for r in rows if commercial(r))
    men = sum(1 for r in rows if ysl_mention(r))
    cit = sum(1 for r in rows
              if (r.get("Reference ID"), r.get("Cycle"), r.get("Channel")) in citation_keys)
    return [
        ("Query Set", qs, 1.0 if qs else 0.0),
        ("AI Existence", ai, safe_div(ai, qs)),
        ("Commercial", com, safe_div(com, qs)),
        ("Mention", men, safe_div(men, qs)),
        ("Citation", cit, safe_div(cit, qs)),
    ]


def build_citation_keys(citation: list[dict[str, Any]], domain_filter) -> set[tuple]:
    """Citation 시트에서 domain_filter(domain)==True 인 행의 (Reference ID, Cycle, Platform) 키 set."""
    keys: set[tuple] = set()
    for r in citation:
        if domain_filter(domain_lower(r.get("Domain"))):
            keys.add((r.get("Reference ID"), r.get("Cycle"), r.get("Platform")))
    return keys


# ---------------------------------------------------------------------------
# 시트 빌더
# ---------------------------------------------------------------------------

@dataclass
class Sheet:
    name: str
    section: str
    headers: list[str]
    rows: list[list[Any]]


SECTION_PERIOD: list[str] = [""]
SECTION_BRAND = "YSL Beauty"
SECTION_SOURCE = ""


def make_sheet(name: str, section: str, headers: list[str], rows: list[list[Any]]) -> Sheet:
    return Sheet(name=name, section=section, headers=headers, rows=rows)


def s01_data_overview(raw: Raw) -> Sheet:
    mention = raw.mention
    citation = raw.citation
    dates = [r.get("Date") for r in mention if r.get("Date")]
    period = f"{min(dates)} ~ {max(dates)}" if dates else "-"
    cycles = sorted({r.get("Cycle") for r in mention if r.get("Cycle") is not None})
    cat_counts = Counter(r.get("Category") for r in raw.query_list if r.get("Category"))
    resp_counts = Counter(r.get("Category") for r in mention if r.get("Category"))

    rows: list[list[Any]] = [
        ["Unique Queries", len(raw.query_list)],
        ["Total Responses", len(mention)],
        ["Total Citations", len(citation)],
        ["Channels", ", ".join(CHANNEL_LABELS[c] for c in CHANNELS)],
        ["Responses Per Query (avg)", round(len(mention) / max(len(raw.query_list), 1), 2)],
        ["Analysis Period", period],
        ["Brand", "YSL Beauty"],
        ["Cycles", ", ".join(str(c) for c in cycles)],
    ]
    for cat in CATEGORIES:
        rows.append([f"Category Mix — {cat} (Questions)", cat_counts.get(cat, 0)])
        rows.append([f"Category Mix — {cat} (Responses)", resp_counts.get(cat, 0)])

    # 자사몰 인용 URL 총 건수 (응답 단위 아닌 URL 단위)
    own_url_count = sum(1 for r in citation if "yslbeauty" in domain_lower(r.get("Domain")))
    rows.append(["YSL Own-Site Citation URL count (total)", own_url_count])

    return make_sheet("01_Data_Overview", "1-1. Data Overview",
                      ["Metric", "Value"], rows)


def s02_position_intent(raw: Raw) -> Sheet:
    """Position(Axis) × Intent Type Value × {논브랜드 응답수, 논브랜드 쿼리수, 브랜드 응답수, 브랜드 쿼리수}"""
    # Reference ID → Axis values
    qmap = {r["Reference ID (Q)"]: r for r in raw.query_list if r.get("Reference ID (Q)")}
    kmap = {r["Reference ID (KW)"]: r for r in raw.query_list if r.get("Reference ID (KW)")}

    def get_axis_value(ref_id: str, axis_col: str) -> str | None:
        rec = qmap.get(ref_id) or kmap.get(ref_id)
        if not rec:
            return None
        v = rec.get(axis_col)
        if v is None or str(v).strip() in ("-", "", "None"):
            return None
        return str(v)

    rows: list[list[Any]] = []
    grand = {"nb_resp": 0, "nb_q": set(), "br_resp": 0, "br_q": set()}

    for axis_no, axis_label, axis_col in AXIS_LABELS:
        axis_total = {"nb_resp": 0, "nb_q": set(), "br_resp": 0, "br_q": set()}
        # value별 카운트
        per_value: dict[str, dict[str, Any]] = defaultdict(
            lambda: {"nb_resp": 0, "nb_q": set(), "br_resp": 0, "br_q": set()})
        for r in raw.mention:
            ref = r.get("Reference ID")
            if not ref:
                continue
            v = get_axis_value(ref, axis_col)
            if not v:
                continue
            is_brand = str(r.get("Brand / Non-Brand", "")).strip() == "Brand"
            slot = per_value[v]
            if is_brand:
                slot["br_resp"] += 1
                slot["br_q"].add(ref)
                axis_total["br_resp"] += 1
                axis_total["br_q"].add(ref)
                grand["br_resp"] += 1
                grand["br_q"].add(ref)
            else:
                slot["nb_resp"] += 1
                slot["nb_q"].add(ref)
                axis_total["nb_resp"] += 1
                axis_total["nb_q"].add(ref)
                grand["nb_resp"] += 1
                grand["nb_q"].add(ref)

        # 출력
        for v in sorted(per_value.keys()):
            d = per_value[v]
            rows.append([f"{axis_no} ({axis_label})", v,
                         d["nb_resp"], len(d["nb_q"]), d["br_resp"], len(d["br_q"])])
        rows.append([f"{axis_no} ({axis_label})", "Total",
                     axis_total["nb_resp"], len(axis_total["nb_q"]),
                     axis_total["br_resp"], len(axis_total["br_q"])])

    rows.append(["Grand Total", "",
                 grand["nb_resp"], len(grand["nb_q"]),
                 grand["br_resp"], len(grand["br_q"])])

    return make_sheet("02_Position_Intent", "1-2. Position and Intent Type",
                      ["Position (Axis)", "Intent Type Value",
                       "논브랜드 응답수", "논브랜드 쿼리수",
                       "브랜드 응답수", "브랜드 쿼리수"], rows)


def s03_category(raw: Raw) -> Sheet:
    """Category × {목적, 질문수, 키워드수, 응답수, 비율}"""
    q_count = Counter(r.get("Category") for r in raw.query_list if r.get("Category"))
    kw_count = Counter()
    for r in raw.query_list:
        if r.get("Category") and r.get("Reference ID (KW)"):
            kw_count[r["Category"]] += 1
    resp_count = Counter(r.get("Category") for r in raw.mention if r.get("Category"))
    total_q = sum(q_count.values()) or 1

    rows: list[list[Any]] = []
    for cat in CATEGORIES:
        purpose = raw.category_purpose.get(cat, "")
        rows.append([cat, purpose,
                     q_count.get(cat, 0),
                     kw_count.get(cat, 0),
                     resp_count.get(cat, 0),
                     round(q_count.get(cat, 0) / total_q, 4)])
    rows.append(["Grand Total", "",
                 sum(q_count.values()), sum(kw_count.values()),
                 sum(resp_count.values()), 1.0])

    return make_sheet("03_Category", "1-3. Category",
                      ["Category", "목적", "질문 수", "키워드 수", "응답 수", "비율"], rows)


def _funnel_rows_for(rows: list[dict[str, Any]], citation_keys: set[tuple]):
    return compute_funnel(rows, citation_keys)


def _ysl_citation_keys(citation: list[dict[str, Any]]) -> set[tuple]:
    return build_citation_keys(citation, lambda d: "yslbeauty" in d)


def s04_overall_funnel(raw: Raw) -> Sheet:
    keys = _ysl_citation_keys(raw.citation)
    funnel = _funnel_rows_for(raw.mention, keys)
    rows = [[stage, c, rate] for (stage, c, rate) in funnel]
    return make_sheet("04_Overall_Funnel", "2-1 A. Overall Funnel",
                      ["Stage", "Count", "Rate"], rows)


def s05_channel_funnel(raw: Raw) -> Sheet:
    keys = _ysl_citation_keys(raw.citation)
    rows: list[list[Any]] = []
    for ch in CHANNELS:
        sub = [r for r in raw.mention if r.get("Channel") == ch]
        for stage, c, rate in compute_funnel(sub, keys):
            rows.append([CHANNEL_LABELS[ch], stage, c, rate])
    return make_sheet("05_Channel_Funnel", "2-1 B. Channel Funnel",
                      ["Channel", "Stage", "Count", "Rate"], rows)


def s06_category_funnel(raw: Raw) -> Sheet:
    keys = _ysl_citation_keys(raw.citation)
    rows: list[list[Any]] = []
    for cat in CATEGORIES:
        sub = [r for r in raw.mention if r.get("Category") == cat]
        for stage, c, rate in compute_funnel(sub, keys):
            rows.append([cat, stage, c, rate])
    return make_sheet("06_Category_Funnel", "2-1 C. Category Funnel",
                      ["Category", "Stage", "Count", "Rate"], rows)


def s07_intent_funnel(raw: Raw) -> Sheet:
    keys = _ysl_citation_keys(raw.citation)
    rows: list[list[Any]] = []
    qmap = {r["Reference ID (Q)"]: r for r in raw.query_list if r.get("Reference ID (Q)")}
    kmap = {r["Reference ID (KW)"]: r for r in raw.query_list if r.get("Reference ID (KW)")}
    # mention 에는 Intent 컬럼이 없으므로 Reference ID join
    for intent in INTENTS:
        sub = []
        for r in raw.mention:
            ref = r.get("Reference ID")
            rec = qmap.get(ref) or kmap.get(ref)
            if rec and rec.get("Intent") == intent:
                sub.append(r)
        for stage, c, rate in compute_funnel(sub, keys):
            rows.append([intent, stage, c, rate])
    return make_sheet("07_Intent_Funnel", "2-1 D. Intent Funnel",
                      ["Intent", "Stage", "Count", "Rate"], rows)


def _competitor_matrix(rows: list[dict[str, Any]]) -> list[list[Any]]:
    """AI Engine × Brand 매트릭스 (long-format)."""
    out: list[list[Any]] = []
    for ch in CHANNELS:
        sub = [r for r in rows if r.get("Channel") == ch]
        questions = len(sub)
        total_brand_men = sum(1 for r in sub if commercial(r))
        for b in BRANDS:
            cnt = sum(1 for r in sub if is_yes(r.get(f"{b} 언급")))
            out.append([CHANNEL_LABELS[ch], questions, total_brand_men, b, cnt, safe_div(cnt, questions)])
        out.append([CHANNEL_LABELS[ch], questions, total_brand_men, "Grand Total", total_brand_men, safe_div(total_brand_men, questions)])
    # 전체
    questions = len(rows)
    total_brand_men = sum(1 for r in rows if commercial(r))
    for b in BRANDS:
        cnt = sum(1 for r in rows if is_yes(r.get(f"{b} 언급")))
        out.append(["All", questions, total_brand_men, b, cnt, safe_div(cnt, questions)])
    out.append(["All", questions, total_brand_men, "Grand Total", total_brand_men, safe_div(total_brand_men, questions)])
    return out


def s08_mention_by_competitors(raw: Raw) -> Sheet:
    rows = _competitor_matrix(raw.mention)
    return make_sheet("08_Mention_by_Competitors", "2-2. Overall Mention Rate by Competitors",
                      ["AI Engine", "Questions", "Total Brand Mentions", "Brand", "Mention Count", "Rate"], rows)


def s09_citation_ysl_domain(raw: Raw) -> Sheet:
    """YSL 자사 도메인 인용 (응답 unique 가 아니라 URL 단위 카운트)."""
    counter: Counter = Counter()
    for r in raw.citation:
        d = domain_lower(r.get("Domain"))
        if "yslbeauty" in d:
            counter[(str(r.get("Platform", "")).strip(), r.get("Domain"), r.get("URL"))] += 1
    rows: list[list[Any]] = []
    for i, ((platform, domain, url), c) in enumerate(counter.most_common(), start=1):
        rows.append([i, CHANNEL_LABELS.get(platform, platform), domain, url, c])
    return make_sheet("09_Citation_YSL_Domain", "2-3 A. 자사 도메인 인용",
                      ["No.", "AI Platform", "Domain", "URL", "인용 수"], rows)


def s10_citation_ecommerce(raw: Raw) -> Sheet:
    """E-commerce 채널 인용 (자사몰 우선 매칭 후 E-commerce). 응답에 YSL 언급이 있던 인용만 표기."""
    # 응답 (Reference ID, Cycle, Channel) → YSL 언급 여부
    ysl_resp_keys = {(r.get("Reference ID"), r.get("Cycle"), r.get("Channel"))
                     for r in raw.mention if ysl_mention(r)}
    counter: Counter = Counter()
    for r in raw.citation:
        d = domain_lower(r.get("Domain"))
        if any(kw in d for kw in BRAND_OWN_DOMAIN_KEYWORDS["YSL Beauty"]):
            continue  # 자사몰은 9번 시트로
        retailer = ecommerce_retailer(d)
        if not retailer:
            continue
        url = str(r.get("URL", ""))
        is_ysl_related = (r.get("Reference ID"), r.get("Cycle"), r.get("Platform")) in ysl_resp_keys
        ptype = page_type(url)
        counter[(CHANNEL_LABELS.get(str(r.get("Platform", "")), str(r.get("Platform", ""))),
                 retailer, url, ptype, "Y" if is_ysl_related else "N")] += 1
    rows: list[list[Any]] = []
    for i, ((platform, retailer, url, ptype, ysl_rel), c) in enumerate(counter.most_common(), start=1):
        rows.append([i, platform, retailer, url, ptype, ysl_rel, c])
    return make_sheet("10_Citation_Ecommerce", "2-3 B. E-Commerce Channel 인용",
                      ["No.", "AI Platform", "Channel (Retailer)", "Target URL",
                       "페이지 유형", "YSL 언급 응답 (Y/N)", "인용 수"], rows)


def s11_citation_domain_rank(raw: Raw) -> Sheet:
    """Citation Domain 순위 (Platform × Domain × Domain Type)."""
    counter: Counter = Counter()
    total_per_platform: Counter = Counter()
    for r in raw.citation:
        platform = str(r.get("Platform", ""))
        total_per_platform[platform] += 1
        counter[(platform, r.get("Domain"), r.get("Domain Type"))] += 1
    rows: list[list[Any]] = []
    # Platform별 Top 30
    rank: dict[str, int] = defaultdict(int)
    for (platform, domain, dtype), c in counter.most_common():
        rank[platform] += 1
        if rank[platform] > 30:
            continue
        total = total_per_platform[platform] or 1
        rows.append([rank[platform], CHANNEL_LABELS.get(platform, platform),
                     domain, dtype, c, c / total])
    return make_sheet("11_Citation_Domain_Rank", "2-4 A. Citation Chart Table (도메인 순위)",
                      ["Rank", "AI Platform", "Domain", "Domain Type", "Number of Citation", "Citation Rate"], rows)


def s12_citation_domain_type_share(raw: Raw) -> Sheet:
    """Platform별 Domain Type 분포 (raw 10개 분류 그대로)."""
    counter: Counter = Counter()
    total: Counter = Counter()
    for r in raw.citation:
        platform = str(r.get("Platform", ""))
        dtype = str(r.get("Domain Type", "")) or "(unknown)"
        counter[(platform, dtype)] += 1
        total[platform] += 1
    rows: list[list[Any]] = []
    for ch in CHANNELS:
        platform_total = total[ch] or 1
        for (platform, dtype), c in sorted(counter.items()):
            if platform == ch:
                rows.append([CHANNEL_LABELS[ch], dtype, c, c / platform_total])
    # Total
    grand = sum(total.values()) or 1
    type_total: Counter = Counter()
    for (_, dtype), c in counter.items():
        type_total[dtype] += c
    for dtype in sorted(type_total.keys()):
        rows.append(["Total", dtype, type_total[dtype], type_total[dtype] / grand])
    return make_sheet("12_Citation_DomainType_Share", "2-4 B. Citation Bar Table (도메인 타입 분포)",
                      ["AI Platform", "Domain Type", "Citation Count", "Share"], rows)


def s13_topic_funnel_overall(raw: Raw) -> Sheet:
    keys = _ysl_citation_keys(raw.citation)
    rows: list[list[Any]] = []
    for cat in CATEGORIES:
        sub = [r for r in raw.mention if r.get("Category") == cat]
        for stage, c, rate in compute_funnel(sub, keys):
            rows.append([cat, stage, c, rate])
    return make_sheet("13_Topic_Funnel_Overall", "3-1 A. Topic Overall Funnel",
                      ["Topic", "Stage", "Count", "Rate"], rows)


def s14_topic_funnel_channel(raw: Raw) -> Sheet:
    keys = _ysl_citation_keys(raw.citation)
    rows: list[list[Any]] = []
    for cat in CATEGORIES:
        for ch in CHANNELS:
            sub = [r for r in raw.mention if r.get("Category") == cat and r.get("Channel") == ch]
            for stage, c, rate in compute_funnel(sub, keys):
                rows.append([cat, CHANNEL_LABELS[ch], stage, c, rate])
    return make_sheet("14_Topic_Funnel_Channel", "3-1 B. Topic Channel Funnel",
                      ["Topic", "Channel", "Stage", "Count", "Rate"], rows)


def s15_topic_funnel_intent(raw: Raw) -> Sheet:
    keys = _ysl_citation_keys(raw.citation)
    qmap = {r["Reference ID (Q)"]: r for r in raw.query_list if r.get("Reference ID (Q)")}
    kmap = {r["Reference ID (KW)"]: r for r in raw.query_list if r.get("Reference ID (KW)")}
    rows: list[list[Any]] = []
    for cat in CATEGORIES:
        for intent in INTENTS:
            sub = []
            for r in raw.mention:
                if r.get("Category") != cat:
                    continue
                ref = r.get("Reference ID")
                rec = qmap.get(ref) or kmap.get(ref)
                if rec and rec.get("Intent") == intent:
                    sub.append(r)
            for stage, c, rate in compute_funnel(sub, keys):
                rows.append([cat, intent, stage, c, rate])
    return make_sheet("15_Topic_Funnel_Intent", "3-1 C. Topic Intent Funnel",
                      ["Topic", "Intent", "Stage", "Count", "Rate"], rows)


def s16_topic_mention_by_competitors(raw: Raw) -> Sheet:
    rows: list[list[Any]] = []
    for cat in CATEGORIES:
        sub = [r for r in raw.mention if r.get("Category") == cat]
        matrix = _competitor_matrix(sub)
        for line in matrix:
            rows.append([cat] + line)
    return make_sheet("16_Topic_Mention_by_Competitors", "3-2. Topic Mention Rate by Competitors",
                      ["Topic", "AI Engine", "Questions", "Total Brand Mentions",
                       "Brand", "Mention Count", "Rate"], rows)


def s17_topic_customer_journey(raw: Raw) -> Sheet:
    """Topic × Intent × Brand × Mention Rate (구매여정 4단계)."""
    qmap = {r["Reference ID (Q)"]: r for r in raw.query_list if r.get("Reference ID (Q)")}
    kmap = {r["Reference ID (KW)"]: r for r in raw.query_list if r.get("Reference ID (KW)")}
    rows: list[list[Any]] = []
    for cat in CATEGORIES:
        for intent in INTENTS:
            sub = []
            for r in raw.mention:
                if r.get("Category") != cat:
                    continue
                ref = r.get("Reference ID")
                rec = qmap.get(ref) or kmap.get(ref)
                if rec and rec.get("Intent") == intent:
                    sub.append(r)
            denom = len(sub) or 1
            for b in BRANDS:
                cnt = sum(1 for r in sub if is_yes(r.get(f"{b} 언급")))
                rows.append([cat, intent, b, cnt, cnt / denom if sub else 0.0])
    return make_sheet("17_Topic_Customer_Journey", "3-3. Topic Mention Rate by Customer Decision Making Journey",
                      ["Topic", "Intent", "Brand", "Mention Count", "Mention Rate"], rows)


def s18_topic_positioning(raw: Raw) -> Sheet:
    """Topic × Position(Axis) × Axis Value × Brand × Mention Rate."""
    qmap = {r["Reference ID (Q)"]: r for r in raw.query_list if r.get("Reference ID (Q)")}
    kmap = {r["Reference ID (KW)"]: r for r in raw.query_list if r.get("Reference ID (KW)")}

    def get_axis_value(ref_id: str, axis_col: str) -> str | None:
        rec = qmap.get(ref_id) or kmap.get(ref_id)
        if not rec:
            return None
        v = rec.get(axis_col)
        if v is None or str(v).strip() in ("-", "", "None"):
            return None
        return str(v)

    rows: list[list[Any]] = []
    for cat in CATEGORIES:
        cat_rows = [r for r in raw.mention if r.get("Category") == cat]
        for axis_no, axis_label, axis_col in AXIS_LABELS:
            buckets: dict[str, list[dict]] = defaultdict(list)
            for r in cat_rows:
                v = get_axis_value(r.get("Reference ID"), axis_col)
                if v:
                    buckets[v].append(r)
            for v in sorted(buckets.keys()):
                sub = buckets[v]
                denom = len(sub) or 1
                for b in BRANDS:
                    cnt = sum(1 for r in sub if is_yes(r.get(f"{b} 언급")))
                    rows.append([cat, f"{axis_no} ({axis_label})", v, b,
                                 len(sub), cnt, cnt / denom])
    return make_sheet("18_Topic_Positioning", "3-4. Positioning Decision Mention Rate",
                      ["Topic", "Position (Axis)", "Axis Value", "Brand",
                       "전체 응답 수", "Brand Mention 수", "Mention Rate"], rows)


def s19_topic_citation_domain_rank(raw: Raw) -> Sheet:
    counter: Counter = Counter()
    total: Counter = Counter()
    for r in raw.citation:
        cat = r.get("Category")
        if cat not in CATEGORIES:
            continue
        platform = str(r.get("Platform", ""))
        total[(cat, platform)] += 1
        counter[(cat, platform, r.get("Domain"), r.get("Domain Type"))] += 1
    rows: list[list[Any]] = []
    rank: dict[tuple, int] = defaultdict(int)
    for (cat, platform, domain, dtype), c in sorted(counter.items(), key=lambda x: (x[0][0], -x[1])):
        key = (cat, platform)
        rank[key] += 1
        if rank[key] > 20:
            continue
        denom = total[key] or 1
        rows.append([cat, rank[key], CHANNEL_LABELS.get(platform, platform),
                     domain, dtype, c, c / denom])
    return make_sheet("19_Topic_Citation_Domain_Rank", "3-5 A. Topic Citation Chart Table",
                      ["Topic", "Rank", "AI Platform", "Domain", "Domain Type",
                       "Number of Citation", "Citation Rate"], rows)


def s20_topic_citation_domain_type_share(raw: Raw) -> Sheet:
    counter: Counter = Counter()
    total: Counter = Counter()
    for r in raw.citation:
        cat = r.get("Category")
        if cat not in CATEGORIES:
            continue
        platform = str(r.get("Platform", ""))
        dtype = str(r.get("Domain Type", "")) or "(unknown)"
        counter[(cat, platform, dtype)] += 1
        total[(cat, platform)] += 1
    rows: list[list[Any]] = []
    for cat in CATEGORIES:
        for ch in CHANNELS:
            denom = total[(cat, ch)] or 1
            for (c2, p2, dtype), cnt in sorted(counter.items()):
                if c2 == cat and p2 == ch:
                    rows.append([cat, CHANNEL_LABELS[ch], dtype, cnt, cnt / denom])
    return make_sheet("20_Topic_Citation_DomainType_Share", "3-5 B. Topic Citation Bar Table",
                      ["Topic", "AI Platform", "Domain Type", "Citation Count", "Share"], rows)


def _topic_top5_by_domain_filter(raw: Raw, name: str, domain_filter, blank_extra: list[str]) -> Sheet:
    counter_url: Counter = Counter()
    title_map: dict[str, str] = {}
    for r in raw.citation:
        cat = r.get("Category")
        if cat not in CATEGORIES:
            continue
        d = domain_lower(r.get("Domain"))
        if not domain_filter(d):
            continue
        url = str(r.get("URL", ""))
        if not url:
            continue
        counter_url[(cat, url)] += 1
        if r.get("Title") and url not in title_map:
            title_map[url] = str(r["Title"])
    rows: list[list[Any]] = []
    per_topic: dict[str, list] = defaultdict(list)
    for (cat, url), c in counter_url.most_common():
        if len(per_topic[cat]) < 5:
            per_topic[cat].append((url, c))
    for cat in CATEGORIES:
        for rank, (url, c) in enumerate(per_topic.get(cat, []), start=1):
            rows.append([cat, rank, title_map.get(url, ""), url, c] + ["" for _ in blank_extra])
    return make_sheet(name, name.replace("_", ".") , ["Topic", "Rank", "Title", "URL", "인용 수"] + blank_extra, rows)


def s21_topic_youtube_top5(raw: Raw) -> Sheet:
    sheet = _topic_top5_by_domain_filter(
        raw, "21_Topic_YouTube_Top5",
        domain_filter=lambda d: "youtube.com" in d,
        blank_extra=["콘텐츠 주제", "소구 메시지 특징"])
    sheet.section = "3-6. Topic YouTube Top 5 Citation"
    return sheet


def s22_topic_blog_top5(raw: Raw) -> Sheet:
    sheet = _topic_top5_by_domain_filter(
        raw, "22_Topic_Blog_Top5",
        domain_filter=lambda d: "blog.naver.com" in d or "tistory.com" in d,
        blank_extra=["콘텐츠 주제", "소구 메시지 특징"])
    sheet.section = "3-6. Topic Blog Top 5 Citation"
    return sheet


def s23_topic_ecommerce_pages(raw: Raw) -> Sheet:
    """Topic × 리테일러 × 페이지 유형별 인용 수 (Wide format으로 출력)."""
    matrix: dict[tuple, int] = defaultdict(int)
    page_types_order = ["검색결과", "상품 상세 (PDP)", "기획전/매거진", "기타"]
    for r in raw.citation:
        cat = r.get("Category")
        if cat not in CATEGORIES:
            continue
        d = domain_lower(r.get("Domain"))
        if any(kw in d for kw in BRAND_OWN_DOMAIN_KEYWORDS["YSL Beauty"]):
            continue
        retailer = ecommerce_retailer(d)
        if not retailer:
            continue
        ptype = page_type(str(r.get("URL", "")))
        matrix[(cat, retailer, ptype)] += 1

    rows: list[list[Any]] = []
    seen_retailers: dict[str, set] = defaultdict(set)
    for (cat, retailer, _), _ in matrix.items():
        seen_retailers[cat].add(retailer)

    for cat in CATEGORIES:
        for retailer in sorted(seen_retailers.get(cat, [])):
            counts = [matrix.get((cat, retailer, pt), 0) for pt in page_types_order]
            rows.append([cat, retailer] + counts + [sum(counts)])
    headers = ["Topic", "리테일러"] + page_types_order + ["합계"]
    return make_sheet("23_Topic_Ecommerce_Pages", "3-7 A. Topic E-Commerce 페이지 유형별 인용",
                      headers, rows)


def s24_topic_ecommerce_summary(raw: Raw) -> Sheet:
    """Topic × 리테일러 × 주요 인용 페이지 유형 / 자사 브랜드 관련 / PDP 예시 URL / 인용 특징(빈)"""
    ysl_resp_keys = {(r.get("Reference ID"), r.get("Cycle"), r.get("Channel"))
                     for r in raw.mention if ysl_mention(r)}
    pcounter: dict[tuple, Counter] = defaultdict(Counter)  # (Topic, Retailer) -> page_type counter
    pdp_example: dict[tuple, str] = {}
    ysl_related: dict[tuple, bool] = defaultdict(bool)
    for r in raw.citation:
        cat = r.get("Category")
        if cat not in CATEGORIES:
            continue
        d = domain_lower(r.get("Domain"))
        if any(kw in d for kw in BRAND_OWN_DOMAIN_KEYWORDS["YSL Beauty"]):
            continue
        retailer = ecommerce_retailer(d)
        if not retailer:
            continue
        url = str(r.get("URL", ""))
        ptype = page_type(url)
        pcounter[(cat, retailer)][ptype] += 1
        if ptype == "상품 상세 (PDP)" and (cat, retailer) not in pdp_example:
            pdp_example[(cat, retailer)] = url
        if (r.get("Reference ID"), r.get("Cycle"), r.get("Platform")) in ysl_resp_keys:
            ysl_related[(cat, retailer)] = True

    rows: list[list[Any]] = []
    for (cat, retailer), pc in sorted(pcounter.items()):
        top_type = pc.most_common(1)[0][0] if pc else ""
        rows.append([cat, retailer, top_type, "", "Y" if ysl_related[(cat, retailer)] else "N",
                     pdp_example.get((cat, retailer), "")])
    return make_sheet("24_Topic_Ecommerce_Summary", "3-7 B. Topic E-Commerce 인용 특징 요약",
                      ["Topic", "리테일러", "주요 인용 페이지 유형", "인용 특징",
                       "자사 브랜드 관련 (Y/N)", "PDP 예시 URL"], rows)


def s25_topic_brand_citation(raw: Raw) -> Sheet:
    """Topic × Brand × {Mention 건수, 자사몰 Citation, Citation Rate, 핵심 인용 콘텐츠 유형(auto), 시사점(빈)}"""
    rows: list[list[Any]] = []
    for cat in CATEGORIES:
        cat_mention = [r for r in raw.mention if r.get("Category") == cat]
        cat_citation = [r for r in raw.citation if r.get("Category") == cat]
        for b in BRANDS:
            mcnt = sum(1 for r in cat_mention if is_yes(r.get(f"{b} 언급")))
            kws = BRAND_OWN_DOMAIN_KEYWORDS.get(b, [])
            own_cits = [r for r in cat_citation
                        if any(kw.lower() in domain_lower(r.get("Domain")) for kw in kws)]
            own_count = len(own_cits)
            crate = own_count / mcnt if mcnt else 0.0
            # 핵심 인용 콘텐츠 유형 = own_cits 의 Domain Type 1위
            dtypes = Counter(str(r.get("Domain Type", "")) for r in own_cits if r.get("Domain Type"))
            top_type = dtypes.most_common(1)[0][0] if dtypes else ""
            rows.append([cat, b, mcnt, own_count, crate, top_type, ""])
    return make_sheet("25_Topic_Brand_Citation", "3-8 A. Topic 브랜드 자사몰 Citation 현황",
                      ["Topic", "Brand", "Mention 건수", "자사몰 Citation",
                       "Citation Rate", "핵심 인용 콘텐츠 유형", "시사점"], rows)


def s26_topic_brand_ownsite_pages(raw: Raw) -> Sheet:
    """Topic × Brand × 페이지 유형별 자사몰 인용."""
    matrix: dict[tuple, int] = defaultdict(int)
    title_map: dict[tuple, tuple[str, str]] = {}
    for r in raw.citation:
        cat = r.get("Category")
        if cat not in CATEGORIES:
            continue
        d = domain_lower(r.get("Domain"))
        url = str(r.get("URL", ""))
        title = str(r.get("Title", "")) if r.get("Title") else ""
        for b, kws in BRAND_OWN_DOMAIN_KEYWORDS.items():
            if any(kw.lower() in d for kw in kws):
                ptype = page_type(url)
                matrix[(cat, b, ptype)] += 1
                key = (cat, b, ptype)
                if key not in title_map:
                    title_map[key] = (title, url)
                break
    rows: list[list[Any]] = []
    # Topic × Brand 별 합계 → 비율
    totals: dict[tuple, int] = defaultdict(int)
    for (cat, b, _), c in matrix.items():
        totals[(cat, b)] += c
    for (cat, b, ptype), c in sorted(matrix.items()):
        denom = totals[(cat, b)] or 1
        title, url = title_map.get((cat, b, ptype), ("", ""))
        rows.append([cat, b, ptype, c, c / denom, title, url])
    return make_sheet("26_Topic_Brand_OwnSite_Pages", "3-8 B. Topic 브랜드 자사몰 페이지 유형 분석",
                      ["Topic", "Brand", "페이지 유형", "인용 수", "비율",
                       "대표 콘텐츠 예시 (Title)", "대표 URL"], rows)


# ---------------------------------------------------------------------------
# xlsx 출력
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="222222")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12, color="222222")


def write_xlsx(out_path: Path, sheets: list[Sheet], raw: Raw) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    period = "-"
    dates = [r.get("Date") for r in raw.mention if r.get("Date")]
    if dates:
        period = f"{min(dates)} ~ {max(dates)}"

    for s in sheets:
        ws = wb.create_sheet(s.name[:31])  # Excel 시트명 31자 제한
        ws.cell(row=1, column=1, value=f"Section: {s.section}").font = SECTION_FONT
        ws.cell(row=2, column=1, value=f"Source: {raw.mention[0].get('Date', '-') if raw.mention else '-'} (raw)")
        ws.cell(row=3, column=1, value=f"Period: {period}")
        ws.cell(row=4, column=1, value=f"Brand: YSL Beauty")
        # 헤더는 6행
        for j, h in enumerate(s.headers, start=1):
            cell = ws.cell(row=6, column=j, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for i, row in enumerate(s.rows, start=7):
            for j, v in enumerate(row, start=1):
                cell = ws.cell(row=i, column=j, value=v)
                if isinstance(v, float) and 0.0 <= v <= 1.0 and j > 1:
                    cell.number_format = "0.00%"
        # 컬럼 너비 자동
        for j in range(1, len(s.headers) + 1):
            ws.column_dimensions[get_column_letter(j)].width = max(
                12, min(50, max(len(str(s.headers[j-1])), *(len(str(r[j-1])) for r in s.rows[:50] if j-1 < len(r))) + 2))
        ws.freeze_panes = "A7"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[write] {out_path} ({len(sheets)} sheets)")


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

BUILDERS = [
    s01_data_overview, s02_position_intent, s03_category,
    s04_overall_funnel, s05_channel_funnel, s06_category_funnel, s07_intent_funnel,
    s08_mention_by_competitors,
    s09_citation_ysl_domain, s10_citation_ecommerce,
    s11_citation_domain_rank, s12_citation_domain_type_share,
    s13_topic_funnel_overall, s14_topic_funnel_channel, s15_topic_funnel_intent,
    s16_topic_mention_by_competitors,
    s17_topic_customer_journey, s18_topic_positioning,
    s19_topic_citation_domain_rank, s20_topic_citation_domain_type_share,
    s21_topic_youtube_top5, s22_topic_blog_top5,
    s23_topic_ecommerce_pages, s24_topic_ecommerce_summary,
    s25_topic_brand_citation, s26_topic_brand_ownsite_pages,
]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--monitoring", default="projects/ysl/report/Bubbleshare_YSL_VIVI_Monitoring_0504.xlsx")
    parser.add_argument("--question-list", default="projects/ysl/report/Bubbleshare_YSL_Question_List_수정.xlsx")
    parser.add_argument("--output", default=None,
                        help="출력 경로. 미지정 시 입력 suffix 미러링")
    args = parser.parse_args(argv)

    monitoring = Path(args.monitoring)
    question_list = Path(args.question_list)
    if args.output:
        out = Path(args.output)
    else:
        m = re.search(r"_(\d{4})\.xlsx$", monitoring.name)
        suffix = f"_{m.group(1)}" if m else ""
        out = Path(f"projects/ysl/report/Bubbleshare_YSL_Audit_Report_Data{suffix}.xlsx")

    raw = load_raw(monitoring, question_list)
    print(f"[stat] Mention rows = {len(raw.mention)}, Citation rows = {len(raw.citation)}, "
          f"Query List rows = {len(raw.query_list)}")
    sheets = [b(raw) for b in BUILDERS]
    write_xlsx(out, sheets, raw)
    return 0


if __name__ == "__main__":
    sys.exit(main())
