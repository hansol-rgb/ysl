#!/usr/bin/env python3
"""YSL AI Visibility Report v2 빌더.

인풋: final/Bubbleshare_YSL_VIVI_Monitoring*.xlsx + template/brands_config.json
아웃풋: final/Bubbleshare_YSL_AI_Visibility_Report_v2.xlsx (10시트)

시트 구성:
  1. Analysis Background (공통)
  2-4. 합친 버전 (Q+KW): Funnel / Competitive / Content
  5-7. Q-side만 (ChatGPT 1,440)
  8-10. KW-side만 (Google + Naver 2,880)
"""
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
# VIVI Monitoring 파일명에 차수 suffix(`_4차.xlsx`)/날짜 suffix(`_0430.xlsx`)가 섞여 있어
# 알파벳 정렬은 신뢰할 수 없음 → 수정 시간(mtime) 기준 가장 최근 파일 선택
_vivi_candidates = list((ROOT / "final").glob("Bubbleshare_YSL_VIVI_Monitoring*.xlsx"))
SRC_XLSX = max(_vivi_candidates, key=lambda p: p.stat().st_mtime) if _vivi_candidates else ROOT / "final" / "Bubbleshare_YSL_VIVI_Monitoring.xlsx"
BRANDS_JSON = ROOT / "template" / "brands_config.json"
# 출력 파일명: 인풋 VIVI 파일의 suffix(`_0430` / `_9차`)를 그대로 미러링
_vivi_suffix = SRC_XLSX.stem.replace("Bubbleshare_YSL_VIVI_Monitoring", "")
OUT_XLSX = ROOT / "final" / f"Bubbleshare_YSL_AI_Visibility_Report{_vivi_suffix}.xlsx"

# 색상
PRIMARY = "8E2896"
SECONDARY = "752577"
LIGHT = "F3E5F5"
WHITE = "FFFFFF"

CHANNEL_DISPLAY = {
    "google": "Google (AIO)",
    "naver": "Naver",
    "chatgpt": "ChatGPT",
    "perplexity": "Perplexity",
}


# ===========================
# 1. 데이터 로딩
# ===========================
def load_data():
    wb = openpyxl.load_workbook(SRC_XLSX, read_only=True, data_only=True)

    # 02. Mention → responses
    ws = wb["02. Mention"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    responses = [dict(zip(headers, r)) for r in rows[1:] if r[0] is not None]

    # 03. Citation → citations
    ws = wb["03. Citation"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    citations = [dict(zip(headers, r)) for r in rows[1:] if r[0] is not None]

    # 01. Query List → prompts (for Background)
    ws = wb["01. Query List"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    prompts = [dict(zip(headers, r)) for r in rows[1:] if r[0] is not None]

    wb.close()

    with open(BRANDS_JSON, encoding="utf-8") as f:
        brands_config = json.load(f)

    return responses, citations, prompts, brands_config


# ===========================
# 2. Domain Type / Content Type 휴리스틱 분류
# ===========================
DOMAIN_TYPE_RULES = [
    # (label, keywords)
    ("공식몰", [
        "yslbeauty", "dior.com", "chanel.com", "lancome", "esteelauder", "esteelauder.co.kr",
        "jomalone", "narscosmetics", "tomford", "maccosmetics", "hera.com", "sulwhasoo",
        "prada.com", "pradabeauty", "olens.co.kr",
    ]),
    ("리테일/이커머스", [
        "oliveyoung", "ssg.com", "lottedfs", "lotteon", "shinsegae", "hyundaifs",
        "coupang", "gmarket", "11st", "auction", "wadiz.kr", "kakaomakers",
        "smartstore.naver", "shopping.naver", "musinsa", "29cm", "wconcept",
    ]),
    ("뉴스", [
        "khan.co.kr", "joongang.co.kr", "donga.com", "chosun.com", "hankyung.com",
        "mk.co.kr", "mt.co.kr", "sedaily.com", "ytn.co.kr", "kbs.co.kr", "sbs.co.kr",
        "mbc.co.kr", "jtbc.co.kr", "news1.kr", "newsis.com", "yna.co.kr", "yonhapnews",
        "fashionbiz", "fashionn", "newspim", "asiae.co.kr", "edaily.co.kr",
        "biz.heraldcorp", "heraldcorp", "news.mt", "etnews.com", "weeklytrade",
        "fashionpost", "tvchosun", "ohmynews", "munhwa.com", "segye.com",
        "newdaily.co.kr", "noeulnews", "zdnet.co.kr",
    ]),
    ("블로그", [
        "blog.naver.com", "m.blog.naver.com", "tistory.com", "brunch.co.kr",
        "blogspot.com", "medium.com", "cafe.naver.com",  # 카페도 블로그성
    ]),
    ("커뮤니티", [
        "theqoo.net", "ppomppu.co.kr", "ruliweb.com", "dcinside.com", "instiz",
        "pann.nate.com", "mlbpark.donga", "fmkorea.com", "bobaedream.co.kr",
        "todayhumor", "missycoupons", "iladiestoday", "soomgo", "joonggo",
        "powderroom", "82cook", "everytime", "natepann", "ohou.se",
    ]),
    ("영상", [
        "youtube.com", "youtu.be", "vimeo.com", "tiktok.com", "dailymotion",
        "vlive.tv", "afreecatv",
    ]),
    ("SNS", [
        "instagram.com", "twitter.com", "x.com", "facebook.com", "threads.net",
        "weibo.com", "pinterest.com",
    ]),
    ("리뷰/뷰티", [
        "glowpick.com", "hwahae.co.kr", "powderroom.co.kr",
    ]),
    ("위키/지식", [
        "wikipedia.org", "namu.wiki", "namuwiki", "ko.wikipedia",
    ]),
    ("AI/검색", [
        "openai.com", "chatgpt", "perplexity", "gemini.google", "claude.ai",
        "bard.google",
    ]),
    ("매거진", [
        "vogue.co.kr", "elle.co.kr", "marieclaire.kr", "harpersbazaar.co.kr",
        "wkorea.com", "allurekorea", "cosmopolitan.co.kr", "instyle.co.kr",
        "beautypost", "jentestore", "beautynury", "thecelebritymagazine",
    ]),
]

CONTENT_TYPE_KEYWORDS = [
    ("리뷰", ["리뷰", "후기", "써본", "사용후기", "사용기", "review", "솔직"]),
    ("비교", [" vs ", "vs.", "비교", "차이", "vs ", "다른점"]),
    ("추천/큐레이션", ["추천", "베스트", "best ", "top", "랭킹", "순위", "엄선"]),
    ("가이드/팁", ["방법", "가이드", "사용법", "어떻게", "how to", "tips", "꿀팁", "tip"]),
    ("뉴스/이슈", ["발표", "출시", "공개", "런칭", "신상", "news"]),
    ("브랜드 소개", ["브랜드", "이야기", "역사", "스토리", "story"]),
]


# BackOffice 영문 라벨 → 한국어 라벨 매핑 (휴리스틱 라벨과 통일)
BACKOFFICE_DOMAIN_TYPE_MAP = {
    'official': '공식몰',
    'official_blog': '공식 블로그',
    'ecommerce': '리테일/이커머스',
    'news': '뉴스',
    'external_blog': '블로그',
    'forum': '커뮤니티',
    'video': '영상',
    'social_media': 'SNS',
    'wiki': '위키/지식',
    'others': '기타',
}

BACKOFFICE_CONTENT_TYPE_MAP = {
    'curation': '추천/큐레이션',
    'news': '뉴스/이슈',
    'pdp': '상품 페이지',
    'plp': '상품 리스트',
    'tutorial': '가이드/팁',
    'fact': '정보/팩트',
    'review': '리뷰',
    'others': '기타',
    'analysis/insight': '분석/인사이트',
    'qna': 'Q&A',
    'forum': '커뮤니티 글',
    'homepage': '공식 페이지',
    'notice': '공지',
}


def classify_domain_type(domain, url, backoffice_value=None):
    """BackOffice 분류값 우선 + 없으면 도메인 + URL 휴리스틱 fallback."""
    if backoffice_value:
        bo = str(backoffice_value).strip().lower()
        if bo in BACKOFFICE_DOMAIN_TYPE_MAP:
            return BACKOFFICE_DOMAIN_TYPE_MAP[bo]
    if not domain:
        return "기타"
    target = (domain + " " + (url or "")).lower()
    for label, keys in DOMAIN_TYPE_RULES:
        for k in keys:
            if k in target:
                return label
    return "기타"


def classify_content_type(title, url, domain_type, backoffice_value=None):
    """BackOffice 분류값 우선 + 없으면 제목 + URL + 도메인타입 휴리스틱 fallback."""
    if backoffice_value:
        bo = str(backoffice_value).strip().lower()
        if bo in BACKOFFICE_CONTENT_TYPE_MAP:
            return BACKOFFICE_CONTENT_TYPE_MAP[bo]
    if domain_type == "공식몰":
        return "공식 페이지"
    if domain_type == "영상":
        return "영상 콘텐츠"
    if domain_type == "위키/지식":
        return "위키/사전"
    if domain_type == "SNS":
        return "SNS 포스트"

    text = ((title or "") + " " + (url or "")).lower()
    for label, kws in CONTENT_TYPE_KEYWORDS:
        for k in kws:
            if k in text:
                return label
    if domain_type == "뉴스":
        return "뉴스/이슈"
    if domain_type == "커뮤니티":
        return "커뮤니티 글"
    if domain_type == "블로그":
        return "블로그 글"
    if domain_type == "매거진":
        return "매거진 기사"
    if domain_type == "리테일/이커머스":
        return "상품 페이지"
    return "기타"


# ===========================
# 3. 응답 행에 플래그 컴퓨팅
# ===========================
def compute_flags(responses, brands_config):
    brand = brands_config["brand"]
    competitors = brands_config["competitors"]
    all_entities = [brand] + competitors
    brand_url_sub = brand["url"].lower()

    for r in responses:
        # is_ai_existence
        r["_ai"] = (r.get("Has AI Overview") == "Yes")
        # 자사 인용 (Source URLs / Source Domains)
        urls = (r.get("Source URLs") or "").lower()
        doms = (r.get("Source Domains") or "").lower()
        r["_brand_cite"] = (brand_url_sub in urls) or (brand_url_sub in doms)
        # 자사 멘션
        r["_brand_mention"] = (r.get("YSL Beauty 언급") == "Y")
        # 경쟁사 멘션 — VIVI 파일에 사전 계산된 컬럼 활용
        r["_competitor_mentions"] = {}
        for ent in all_entities:
            col = f"{ent['name']} 언급"
            r["_competitor_mentions"][ent["name"]] = (r.get(col) == "Y")
        # is_commercial — 어떤 엔티티든 멘션
        r["_commercial"] = any(r["_competitor_mentions"].values())
    return all_entities


# ===========================
# 4. 데이터셋 분리 (Combined / Q / KW)
# ===========================
def split_datasets(responses, citations):
    combined_resp = responses
    q_resp = [r for r in responses if r.get("Keyword/Query") == "question"]
    kw_resp = [r for r in responses if r.get("Keyword/Query") == "keyword"]

    combined_cit = citations
    q_cit = [c for c in citations if c.get("Keyword/Query") == "question"]
    kw_cit = [c for c in citations if c.get("Keyword/Query") == "keyword"]

    return [
        ("Combined", combined_resp, combined_cit),
        ("Q-side", q_resp, q_cit),
        ("KW-side", kw_resp, kw_cit),
    ]


# ===========================
# 5. 메트릭 계산
# ===========================
def funnel_metrics(rows):
    n = len(rows)
    if n == 0:
        return {"Query Set": 0, "AI Existence": 0, "Commercial": 0, "Mention": 0, "Citation": 0}
    return {
        "Query Set": n,
        "AI Existence": sum(1 for r in rows if r["_ai"]),
        "Commercial": sum(1 for r in rows if r["_commercial"]),
        "Mention": sum(1 for r in rows if r["_brand_mention"]),
        "Citation": sum(1 for r in rows if r["_brand_cite"]),
    }


def funnel_by(rows, key):
    groups = defaultdict(list)
    for r in rows:
        v = r.get(key)
        if v is not None:
            groups[v].append(r)
    return {g: funnel_metrics(rs) for g, rs in groups.items()}


def mention_rate_by(rows, group_key, all_entities):
    """엔티티별 mention rate by group."""
    groups = defaultdict(list)
    for r in rows:
        v = r.get(group_key)
        if v is not None:
            groups[v].append(r)
    out = {}
    for g, rs in groups.items():
        total = len(rs)
        out[g] = {}
        for ent in all_entities:
            cnt = sum(1 for r in rs if r["_competitor_mentions"][ent["name"]])
            out[g][ent["name"]] = (cnt / total) if total else 0
    return out


# ===========================
# 6. Excel 빌더 헬퍼
# ===========================
def style_header(cell):
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=PRIMARY)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_section(cell):
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=SECONDARY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_analysis_header(cell):
    cell.font = Font(bold=True, color="000000", size=11)
    cell.fill = PatternFill("solid", fgColor=LIGHT)


def autosize(ws, min_width=12):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if not hasattr(cell, "column_letter") or cell.column_letter is None:
                continue
            v = str(cell.value) if cell.value is not None else ""
            w = sum(2 if ord(c) > 127 else 1 for c in v)
            if w > widths.get(cell.column_letter, 0):
                widths[cell.column_letter] = w
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = max(min_width, min(w + 2, 60))


def write_section_header(ws, row, text, span=2):
    ws.cell(row=row, column=1, value=text)
    style_section(ws.cell(row=row, column=1))
    if span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    return row + 1


def write_table_headers(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        style_header(c)
    return row + 1


# ===========================
# 7. 시트 빌더
# ===========================
def build_background_sheet(wb, responses, prompts):
    ws = wb.create_sheet("01. Analysis Background")
    ws.sheet_properties.tabColor = PRIMARY

    row = 1
    ws.cell(row=row, column=1, value="YSL Beauty AI Visibility Report v2 — Analysis Background")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color=PRIMARY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    # Section A — Data Overview
    row = write_section_header(ws, row, "Section A — Data Overview", span=2)
    overview = [
        ("Unique Queries", len({r["Search Query"] if "Search Query" in r else r.get("Question") for r in responses})),
        ("Total Responses", len(responses)),
        ("Channels", ", ".join(sorted({CHANNEL_DISPLAY.get(r["Channel"], r["Channel"]) for r in responses}))),
        ("Categories", ", ".join(sorted({r["Category"] for r in responses if r.get("Category")}))),
        ("Cycles", ", ".join(str(c) for c in sorted({r["Cycle"] for r in responses if r.get("Cycle")}))),
        ("Responses per Query (avg)", round(len(responses) / max(1, len({r.get("Question") for r in responses})), 1)),
        ("Q-side (ChatGPT)", sum(1 for r in responses if r.get("Keyword/Query") == "question")),
        ("KW-side (Google + Naver)", sum(1 for r in responses if r.get("Keyword/Query") == "keyword")),
    ]
    row = write_table_headers(ws, row, ["Metric", "Value"])
    for k, v in overview:
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=v)
        row += 1
    row += 2

    # Section B — Example Prompts by Category
    row = write_section_header(ws, row, "Section B — Example Prompts by Category", span=3)
    row = write_table_headers(ws, row, ["Category", "Example Prompt", "Channel"])
    by_cat = defaultdict(list)
    for r in responses:
        if r["_commercial"] and r.get("Category") and r.get("Question"):
            by_cat[r["Category"]].append((r["Question"], r["Channel"]))
    for cat in sorted(by_cat.keys()):
        seen = set()
        cnt = 0
        for q, ch in by_cat[cat]:
            if q in seen:
                continue
            seen.add(q)
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=q)
            ws.cell(row=row, column=3, value=CHANNEL_DISPLAY.get(ch, ch))
            row += 1
            cnt += 1
            if cnt >= 2:
                break

    autosize(ws)


def write_funnel_table(ws, row, label, metrics_dict_by_group, group_order=None):
    """metrics_dict_by_group: {group_name: {metric: count}} → count + % 페어 칼럼."""
    if group_order is None:
        group_order = list(metrics_dict_by_group.keys())
    metric_keys = ["Query Set", "AI Existence", "Commercial", "Mention", "Citation"]

    # 헤더
    ws.cell(row=row, column=1, value="Metrics")
    style_header(ws.cell(row=row, column=1))
    col = 2
    for g in group_order:
        c1 = ws.cell(row=row, column=col, value=g)
        c2 = ws.cell(row=row, column=col + 1, value=f"{g} (%)")
        style_header(c1)
        style_header(c2)
        col += 2
    row += 1
    # 데이터
    for mk in metric_keys:
        ws.cell(row=row, column=1, value=mk)
        col = 2
        for g in group_order:
            qs = metrics_dict_by_group[g].get("Query Set", 0)
            v = metrics_dict_by_group[g].get(mk, 0)
            ws.cell(row=row, column=col, value=v)
            pct_cell = ws.cell(row=row, column=col + 1, value=(v / qs) if qs else 0)
            pct_cell.number_format = "0.0%"
            col += 2
        row += 1
    return row


def write_overall_funnel(ws, row, metrics):
    ws.cell(row=row, column=1, value="Metrics")
    ws.cell(row=row, column=2, value="Count")
    ws.cell(row=row, column=3, value="Rate")
    style_header(ws.cell(row=row, column=1))
    style_header(ws.cell(row=row, column=2))
    style_header(ws.cell(row=row, column=3))
    row += 1
    qs = metrics.get("Query Set", 0)
    for k in ["Query Set", "AI Existence", "Commercial", "Mention", "Citation"]:
        v = metrics.get(k, 0)
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=v)
        rate = ws.cell(row=row, column=3, value=(v / qs) if qs else 0)
        rate.number_format = "0.0%"
        row += 1
    return row


def write_analysis_block(ws, row, analysis_text, recs):
    """분석 + 추천."""
    row += 2
    c = ws.cell(row=row, column=1, value="Analysis & Insights")
    style_analysis_header(c)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    for line in analysis_text:
        ws.cell(row=row, column=1, value=line)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 30
        row += 1
    row += 1
    c = ws.cell(row=row, column=1, value="Recommended Actions")
    style_analysis_header(c)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1
    for i, rec in enumerate(recs, 1):
        ws.cell(row=row, column=1, value=f"{i}. {rec}")
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 30
        row += 1
    return row


def build_funnel_sheet(wb, name, responses, dataset_label):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = PRIMARY

    row = 1
    ws.cell(row=row, column=1, value=f"Funnel Overview — {dataset_label}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color=PRIMARY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 2

    # Section A — Overall
    row = write_section_header(ws, row, "Section A — Overall Funnel", span=3)
    overall = funnel_metrics(responses)
    row = write_overall_funnel(ws, row, overall)
    row += 2

    # Section B — Channel Funnel
    row = write_section_header(ws, row, "Section B — Channel Funnel", span=8)
    by_channel = funnel_by(responses, "Channel")
    ch_order = sorted(by_channel.keys(), key=lambda x: -by_channel[x].get("Query Set", 0))
    ch_display = {c: CHANNEL_DISPLAY.get(c, c) for c in ch_order}
    by_channel_display = {ch_display[c]: by_channel[c] for c in ch_order}
    row = write_funnel_table(ws, row, "Channel", by_channel_display)
    row += 2

    # Section C — Category Funnel
    row = write_section_header(ws, row, "Section C — Category Funnel", span=8)
    by_cat = funnel_by(responses, "Category")
    cat_order = sorted(by_cat.keys(), key=lambda x: -by_cat[x].get("Query Set", 0))
    by_cat_ord = {c: by_cat[c] for c in cat_order}
    row = write_funnel_table(ws, row, "Category", by_cat_ord)

    # Analysis
    analysis, recs = generate_funnel_analysis(overall, by_channel_display, by_cat_ord, dataset_label)
    row = write_analysis_block(ws, row, analysis, recs)

    autosize(ws, min_width=14)


def build_competitive_sheet(wb, name, responses, all_entities, dataset_label):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = PRIMARY

    row = 1
    ws.cell(row=row, column=1, value=f"Competitive Landscape — {dataset_label}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color=PRIMARY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 2

    # Section A — Mention Rate by Channel
    row = write_section_header(ws, row, "Section A — Mention Rate by Channel", span=8)
    by_ch = mention_rate_by(responses, "Channel", all_entities)
    channels = sorted(by_ch.keys(), key=lambda x: -sum(by_ch[x].values()))
    ws.cell(row=row, column=1, value="Brand")
    style_header(ws.cell(row=row, column=1))
    for i, ch in enumerate(channels, 2):
        c = ws.cell(row=row, column=i, value=CHANNEL_DISPLAY.get(ch, ch))
        style_header(c)
    row += 1
    for ent in all_entities:
        ws.cell(row=row, column=1, value=ent["name"])
        for i, ch in enumerate(channels, 2):
            c = ws.cell(row=row, column=i, value=by_ch[ch].get(ent["name"], 0))
            c.number_format = "0.0%"
        row += 1
    row += 2

    # Section B — Mention Rate by Category
    row = write_section_header(ws, row, "Section B — Mention Rate by Category", span=8)
    by_cat = mention_rate_by(responses, "Category", all_entities)
    cats = sorted(by_cat.keys(), key=lambda x: -sum(by_cat[x].values()))
    ws.cell(row=row, column=1, value="Brand")
    style_header(ws.cell(row=row, column=1))
    for i, cat in enumerate(cats, 2):
        c = ws.cell(row=row, column=i, value=cat)
        style_header(c)
    row += 1
    for ent in all_entities:
        ws.cell(row=row, column=1, value=ent["name"])
        for i, cat in enumerate(cats, 2):
            c = ws.cell(row=row, column=i, value=by_cat[cat].get(ent["name"], 0))
            c.number_format = "0.0%"
        row += 1

    # Analysis
    analysis, recs = generate_competitive_analysis(by_ch, by_cat, all_entities, dataset_label)
    row = write_analysis_block(ws, row, analysis, recs)

    autosize(ws, min_width=14)


def build_content_sheet(wb, name, citations, brand, dataset_label):
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = PRIMARY

    row = 1
    ws.cell(row=row, column=1, value=f"Content Deep Dive — {dataset_label}")
    ws.cell(row=row, column=1).font = Font(bold=True, size=14, color=PRIMARY)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    brand_url = brand["url"].lower()
    total = len(citations)

    # Section A — Brand URL Citations
    row = write_section_header(ws, row, "Section A — Brand URL Citations", span=3)
    brand_cits = [c for c in citations if brand_url in (c.get("URL") or "").lower()]
    brand_total = len(brand_cits)
    url_counts = Counter(c["URL"] for c in brand_cits)
    row = write_table_headers(ws, row, ["URL", "citation_count", "citation_pct"])
    for url, cnt in url_counts.most_common():
        ws.cell(row=row, column=1, value=url)
        ws.cell(row=row, column=2, value=cnt)
        c = ws.cell(row=row, column=3, value=(cnt / brand_total) if brand_total else 0)
        c.number_format = "0.0%"
        row += 1
    if not brand_cits:
        ws.cell(row=row, column=1, value="(자사 URL 인용 없음)")
        row += 1
    row += 2

    # Section B — Top 10 Cited URLs
    row = write_section_header(ws, row, "Section B — Top 10 Cited URLs", span=3)
    url_all = Counter(c["URL"] for c in citations)
    row = write_table_headers(ws, row, ["URL", "citation_count", "citation_pct"])
    for url, cnt in url_all.most_common(10):
        ws.cell(row=row, column=1, value=url)
        ws.cell(row=row, column=2, value=cnt)
        c = ws.cell(row=row, column=3, value=(cnt / total) if total else 0)
        c.number_format = "0.0%"
        row += 1
    row += 2

    # Section C — Top 10 Cited Domains
    row = write_section_header(ws, row, "Section C — Top 10 Cited Domains", span=3)
    dom_all = Counter(c["Domain"] for c in citations if c.get("Domain"))
    row = write_table_headers(ws, row, ["domain", "citation_count", "citation_pct"])
    for d, cnt in dom_all.most_common(10):
        ws.cell(row=row, column=1, value=d)
        ws.cell(row=row, column=2, value=cnt)
        c = ws.cell(row=row, column=3, value=(cnt / total) if total else 0)
        c.number_format = "0.0%"
        row += 1
    row += 2

    # Section D — Domain Type Breakdown
    row = write_section_header(ws, row, "Section D — Domain Type Breakdown", span=3)
    dt_counts = Counter(c.get("_domain_type", "기타") for c in citations)
    row = write_table_headers(ws, row, ["domain_type", "citation_count", "citation_pct"])
    for dt, cnt in dt_counts.most_common():
        ws.cell(row=row, column=1, value=dt)
        ws.cell(row=row, column=2, value=cnt)
        c = ws.cell(row=row, column=3, value=(cnt / total) if total else 0)
        c.number_format = "0.0%"
        row += 1
    row += 2

    # Section E — Content Type Breakdown
    row = write_section_header(ws, row, "Section E — Content Type Breakdown", span=3)
    ct_counts = Counter(c.get("_content_type", "기타") for c in citations)
    row = write_table_headers(ws, row, ["content_type", "citation_count", "citation_pct"])
    for ct, cnt in ct_counts.most_common():
        ws.cell(row=row, column=1, value=ct)
        ws.cell(row=row, column=2, value=cnt)
        c = ws.cell(row=row, column=3, value=(cnt / total) if total else 0)
        c.number_format = "0.0%"
        row += 1

    # Analysis
    analysis, recs = generate_content_analysis(
        brand_cits, brand_total, url_all, dom_all, dt_counts, ct_counts, total, dataset_label
    )
    row = write_analysis_block(ws, row, analysis, recs)

    autosize(ws, min_width=18)


# ===========================
# 8. 분석 텍스트 생성기
# ===========================
def pct(n, d):
    return (n / d * 100) if d else 0


def generate_funnel_analysis(overall, by_channel, by_cat, label):
    qs = overall["Query Set"]
    ai_pct = pct(overall["AI Existence"], qs)
    com_pct = pct(overall["Commercial"], qs)
    men_pct = pct(overall["Mention"], qs)
    cit_pct = pct(overall["Citation"], qs)

    # 채널 격차
    ch_pct = []
    for ch, m in by_channel.items():
        q = m["Query Set"]
        if q == 0: continue
        ch_pct.append((ch, pct(m["Mention"], q), pct(m["AI Existence"], q), pct(m["Citation"], q)))
    ch_pct.sort(key=lambda x: -x[1])
    best_ch = ch_pct[0] if ch_pct else None
    worst_ch = ch_pct[-1] if ch_pct else None

    # 카테고리 격차
    cat_pct = []
    for cat, m in by_cat.items():
        q = m["Query Set"]
        if q == 0: continue
        cat_pct.append((cat, pct(m["Mention"], q), pct(m["AI Existence"], q)))
    cat_pct.sort(key=lambda x: -x[1])
    best_cat = cat_pct[0] if cat_pct else None
    worst_cat = cat_pct[-1] if cat_pct else None

    analysis = [
        f"전체 {qs:,}개 응답 기준, AI Existence {ai_pct:.1f}% → Commercial {com_pct:.1f}% → Mention {men_pct:.1f}% → Citation {cit_pct:.1f}%로 깔때기가 가파르게 좁아집니다. 가장 큰 누수 구간은 Mention({men_pct:.1f}%) → Citation({cit_pct:.1f}%)로, AI가 자사를 입에 올려도 자사 URL을 출처로 인용하는 경우가 거의 없다는 의미입니다.",
    ]
    if best_ch and worst_ch:
        analysis.append(
            f"채널별로는 {best_ch[0]}가 Mention {best_ch[1]:.1f}%로 가장 높고 {worst_ch[0]}가 {worst_ch[1]:.1f}%로 최하입니다. {worst_ch[0]}는 AI Existence({worst_ch[2]:.1f}%) 자체가 낮아 노출 단계에서 이미 누수가 발생합니다."
        )
    if best_cat and worst_cat and best_cat != worst_cat:
        analysis.append(
            f"카테고리별 격차도 큽니다. {best_cat[0]} Mention {best_cat[1]:.1f}% vs {worst_cat[0]} {worst_cat[1]:.1f}%로 {best_cat[1] - worst_cat[1]:.1f}%p 차이. AI가 어떤 카테고리에서 자사 입장을 더 적극적으로 다루는지 데이터 신호가 분명합니다."
        )

    recs = []
    if best_ch and worst_ch and best_ch != worst_ch:
        recs.append(
            f"{worst_ch[0]} 채널 우선 보강 — Mention {worst_ch[1]:.1f}%, Citation {worst_ch[3]:.1f}%로 노출/인용 모두 부진. 해당 채널이 선호하는 출처(블로그/뉴스/공식)를 분석해 콘텐츠 시드 확보 필요."
        )
    recs.append(
        f"Mention→Citation 격차 해소 — 자사몰 콘텐츠 인용률 {cit_pct:.2f}%. 자사 페이지가 AI 출처로 채택될 만한 깊이·구조(가이드형 콘텐츠, 메타데이터, 권위 시그널)를 갖추는 게 시급."
    )
    if worst_cat and best_cat and worst_cat != best_cat:
        recs.append(
            f"{worst_cat[0]} 카테고리에서의 자사 메시지 강화 — Mention {worst_cat[1]:.1f}%로 다른 카테고리(최대 {best_cat[1]:.1f}%) 대비 약함. 카테고리별 USP 메시지 정비."
        )
    return analysis, recs


def generate_competitive_analysis(by_ch, by_cat, all_entities, label):
    brand_name = all_entities[0]["name"]
    # 평균 mention rate 계산
    ent_avg = {}
    for ent in all_entities:
        rates = [by_ch[ch].get(ent["name"], 0) for ch in by_ch.keys() if ch]
        if rates:
            ent_avg[ent["name"]] = sum(rates) / len(rates)
    sorted_ents = sorted(ent_avg.items(), key=lambda x: -x[1])
    brand_idx = next((i for i, (n, _) in enumerate(sorted_ents) if n == brand_name), -1)
    top_competitor = next((n for n, _ in sorted_ents if n != brand_name), None)
    brand_rate = ent_avg.get(brand_name, 0)
    top_rate = ent_avg.get(top_competitor, 0) if top_competitor else 0

    # 카테고리별 자사 위치
    cat_lead = []
    for cat, rates in by_cat.items():
        sorted_in_cat = sorted(rates.items(), key=lambda x: -x[1])
        rank = next((i for i, (n, _) in enumerate(sorted_in_cat) if n == brand_name), -1) + 1
        cat_lead.append((cat, rank, rates.get(brand_name, 0), sorted_in_cat[0]))

    analysis = [
        f"평균 Mention Rate 기준 자사는 11개 경쟁사 풀에서 {brand_idx + 1}위({brand_rate * 100:.1f}%). 1위 {sorted_ents[0][0]}({sorted_ents[0][1] * 100:.1f}%)와 격차 {(sorted_ents[0][1] - brand_rate) * 100:.1f}%p.",
    ]
    if top_competitor:
        analysis.append(
            f"가장 위협적인 경쟁사는 {top_competitor}({top_rate * 100:.1f}%). 채널/카테고리 양 축 모두에서 상위에 노출되어 자사가 들어갈 빈자리를 차지하고 있습니다."
        )
    if cat_lead:
        worst_cat = max(cat_lead, key=lambda x: x[1])
        analysis.append(
            f"카테고리별 입지를 보면 {worst_cat[0]}에서 자사는 {worst_cat[1]}위, 1위는 {worst_cat[3][0]}({worst_cat[3][1] * 100:.1f}%). 이 카테고리는 경쟁사가 AI 응답의 디폴트 옵션으로 자리잡은 상태."
        )

    recs = []
    if top_competitor:
        recs.append(
            f"{top_competitor} 견제 — 평균 {top_rate * 100:.1f}% vs 자사 {brand_rate * 100:.1f}%. 해당 브랜드가 인용되는 채널·카테고리를 매핑해 자사 메시지로 대체 가능한 영역부터 공략."
        )
    if cat_lead:
        worst_cat = max(cat_lead, key=lambda x: x[1])
        recs.append(
            f"{worst_cat[0]} 카테고리 진입 강화 — {worst_cat[1]}위 머무는 동안 {worst_cat[3][0]}가 시장 점유. 해당 카테고리의 검색 의도 맞춤 콘텐츠 시드 6~12개 우선 투입."
        )
    recs.append(
        "전 채널 통합 SOV 모니터링 체계화 — 사이클별 본 표를 자동 생성해 매월 격차 추이 트래킹. 격차 0.5%p 이상 변화 시 즉각 액션 트리거."
    )
    return analysis, recs


def generate_content_analysis(brand_cits, brand_total, url_all, dom_all, dt_counts, ct_counts, total, label):
    # 자사 인용 비중
    brand_share = pct(brand_total, total)
    top_url = url_all.most_common(1)[0] if url_all else None
    top_dom = dom_all.most_common(1)[0] if dom_all else None
    top_dt = dt_counts.most_common(3)
    top_ct = ct_counts.most_common(3)

    analysis = [
        f"전체 {total:,}개 인용 중 자사 URL은 {brand_total}건({brand_share:.2f}%). AI가 자사 콘텐츠를 출처로 채택하는 비율이 사실상 0에 수렴.",
    ]
    if top_dom:
        analysis.append(
            f"가장 많이 인용된 도메인은 {top_dom[0]}({top_dom[1]}건). 도메인 유형은 {' / '.join(f'{n}({c}건, {pct(c, total):.1f}%)' for n, c in top_dt)} 순으로 분포해 AI가 어떤 종류 출처를 신뢰하는지가 드러납니다."
        )
    if top_ct:
        analysis.append(
            f"콘텐츠 종류별로는 {' / '.join(f'{n}({c}건, {pct(c, total):.1f}%)' for n, c in top_ct)} 순. AI가 선호하는 포맷이 명확하므로 자사 콘텐츠 제작 가이드의 직접 인풋이 됩니다."
        )

    recs = []
    if top_dt:
        top_dt_name = top_dt[0][0]
        recs.append(
            f"{top_dt_name} 형식 콘텐츠 자산 확장 — 전체 인용의 {pct(top_dt[0][1], total):.1f}% 차지. 자사가 직접 이 유형 콘텐츠를 생산하거나, 영향력 있는 {top_dt_name} 도메인과 협업해 자사 메시지를 노출."
        )
    if top_ct:
        top_ct_name = top_ct[0][0]
        recs.append(
            f"'{top_ct_name}' 콘텐츠 시리즈화 — AI가 가장 자주 채택하는 콘텐츠 종류({pct(top_ct[0][1], total):.1f}%). 자사 채널/PR 자산을 이 포맷으로 재가공해 인용 가능한 형태로 발행."
        )
    recs.append(
        f"자사몰 페이지 SEO·E-E-A-T 강화 — 현재 자사 인용 {brand_total}건 수준. 카테고리별 hero 페이지(브랜드 스토리·제품 가이드·시즈널 큐레이션) 정비를 통해 인용 가능 페이지 풀 확대."
    )
    if top_dom and "yslbeauty" not in top_dom[0]:
        recs.append(
            f"외부 매체 게재 전략 — {top_dom[0]}({top_dom[1]}건)가 AI 응답의 핵심 출처. 해당 매체에 자사 인용 가능한 기사·인터뷰·리뷰가 게재되도록 PR 우선순위 조정."
        )
    return analysis, recs


# ===========================
# 9. 메인
# ===========================
def main():
    print("[1/5] 데이터 로딩")
    responses, citations, prompts, brands_config = load_data()
    print(f"  responses: {len(responses)}, citations: {len(citations)}, prompts: {len(prompts)}")

    print("[2/5] Citation 분류 (BackOffice 우선 + 휴리스틱 fallback)")
    bo_dom_used = bo_con_used = 0
    for c in citations:
        bo_dom = c.get("Domain Type")
        bo_con = c.get("Content Type")
        c["_domain_type"] = classify_domain_type(c.get("Domain"), c.get("URL"), bo_dom)
        c["_content_type"] = classify_content_type(c.get("Title"), c.get("URL"), c["_domain_type"], bo_con)
        if bo_dom and str(bo_dom).strip().lower() in BACKOFFICE_DOMAIN_TYPE_MAP:
            bo_dom_used += 1
        if bo_con and str(bo_con).strip().lower() in BACKOFFICE_CONTENT_TYPE_MAP:
            bo_con_used += 1
    print(f"  BackOffice Domain Type 적용: {bo_dom_used:,}/{len(citations):,} ({bo_dom_used/len(citations)*100:.1f}%)")
    print(f"  BackOffice Content Type 적용: {bo_con_used:,}/{len(citations):,} ({bo_con_used/len(citations)*100:.1f}%)")

    print("[3/5] Response 플래그")
    all_entities = compute_flags(responses, brands_config)

    print("[4/5] 데이터셋 분리")
    datasets = split_datasets(responses, citations)
    for name, rs, cs in datasets:
        print(f"  {name}: responses {len(rs)}, citations {len(cs)}")

    print("[5/5] Excel 빌드")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: Background
    build_background_sheet(wb, responses, prompts)

    # Sheets 2-10: per dataset
    for ds_name, rs, cs in datasets:
        prefix = {"Combined": "02", "Q-side": "05", "KW-side": "08"}[ds_name]
        order_funnel = {"Combined": 2, "Q-side": 5, "KW-side": 8}[ds_name]
        order_comp = order_funnel + 1
        order_cont = order_funnel + 2
        build_funnel_sheet(wb, f"{order_funnel:02d}. Funnel ({ds_name})", rs, ds_name)
        build_competitive_sheet(wb, f"{order_comp:02d}. Competitive ({ds_name})", rs, all_entities, ds_name)
        build_content_sheet(wb, f"{order_cont:02d}. Content ({ds_name})", cs, brands_config["brand"], ds_name)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"\n✓ Saved: {OUT_XLSX}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    for s in wb.sheetnames:
        print(f"    - {s}")


if __name__ == "__main__":
    main()
