#!/usr/bin/env python3
"""
Audit Report HTML Mockup Builder

`Bubbleshare_YSL_Audit_Report_Data_{suffix}.xlsx` 의 26 시트 + schema 메타를
한 HTML 파일로 변환해서 PPT 슬라이드 미리보기처럼 볼 수 있게 한다.

산출물: `projects/ysl/report/audit_report_mockup_{suffix}.html`

특징:
- 좌측 사이드바 = 26 시트 네비게이션
- 메인 영역 = 슬라이드 1장당 1 섹션 (제목 / PDF 섹션 / 목적 / 표)
- 표는 wide-format 변환 시도 (가능한 시트만), 안 되면 long-format 그대로
- 인쇄/PDF 변환 가능, 키보드 네비 (j/k 다음/이전)
"""

from __future__ import annotations

import argparse
import html
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from audit_report_schema import SHEETS, SheetSpec  # noqa: E402


# ---------------------------------------------------------------------------
# CSS
# ---------------------------------------------------------------------------

CSS = """
* { box-sizing: border-box; }
html, body { margin: 0; padding: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Apple SD Gothic Neo", "Noto Sans KR", sans-serif; }
body { display: flex; min-height: 100vh; background: #f5f5f7; color: #1d1d1f; }
nav.sidebar {
  width: 320px; background: #1d1d1f; color: #f5f5f7; padding: 24px 16px;
  position: sticky; top: 0; height: 100vh; overflow-y: auto;
  font-size: 12.5px;
}
nav.sidebar h1 { font-size: 16px; margin: 0 0 6px; color: #ffffff; }
nav.sidebar .meta { color: #a1a1a6; font-size: 11px; margin-bottom: 16px; }
nav.sidebar a {
  display: block; color: #d2d2d7; text-decoration: none; padding: 6px 8px;
  border-radius: 6px; margin: 2px 0; line-height: 1.35;
}
nav.sidebar a:hover { background: #2c2c2e; color: #ffffff; }
nav.sidebar a.section-head { color: #ff9f0a; font-weight: 600; margin-top: 14px; padding-left: 4px; pointer-events: none; }
nav.sidebar a .num { color: #86868b; font-variant-numeric: tabular-nums; margin-right: 6px; }
main { flex: 1; padding: 40px 56px 120px; max-width: 1200px; }
section.slide {
  background: white; padding: 32px 36px; margin-bottom: 36px;
  border-radius: 12px; box-shadow: 0 1px 3px rgba(0,0,0,0.06);
  scroll-margin-top: 24px;
}
section.slide h2 { margin: 0 0 4px; font-size: 22px; color: #1d1d1f; }
section.slide .pdf-meta { color: #6e6e73; font-size: 12.5px; margin-bottom: 16px; }
section.slide .pdf-meta .pill { display: inline-block; background: #e8f0fe; color: #1a73e8; padding: 2px 8px; border-radius: 12px; font-weight: 600; font-size: 11.5px; margin-right: 6px; }
section.slide .purpose { background: #fafafa; padding: 12px 14px; border-left: 3px solid #ff9f0a; color: #424245; font-size: 13px; margin-bottom: 18px; line-height: 1.5; }
section.slide .data-rules { background: #f4f4f7; padding: 12px 14px; border-radius: 6px; color: #515154; font-size: 12px; margin-bottom: 18px; white-space: pre-wrap; line-height: 1.5; }
section.slide .data-rules strong { color: #1d1d1f; }
table { width: 100%; border-collapse: collapse; font-size: 12.5px; margin-top: 4px; }
table thead th { background: #1d1d1f; color: white; padding: 8px 10px; text-align: left; font-weight: 600; position: sticky; top: 0; }
table tbody td { padding: 6px 10px; border-bottom: 1px solid #e5e5ea; }
table tbody tr:nth-child(even) { background: #fafafa; }
table tbody tr.total { background: #fff8e1 !important; font-weight: 600; }
table tbody td.num { text-align: right; font-variant-numeric: tabular-nums; }
table tbody td.url { font-size: 11px; color: #1a73e8; word-break: break-all; max-width: 380px; }
table tbody td.url a { color: inherit; text-decoration: none; }
.table-wrap { max-height: 480px; overflow: auto; border: 1px solid #e5e5ea; border-radius: 6px; }
.row-count { color: #86868b; font-size: 11px; margin-top: 8px; }
.validation { color: #1a73e8; font-size: 11.5px; margin-top: 12px; padding: 6px 10px; background: #e8f0fe; border-radius: 6px; }
.section-divider {
  margin: 60px 0 28px; padding-top: 28px; border-top: 2px solid #d2d2d7;
  font-size: 14px; color: #86868b; text-transform: uppercase; letter-spacing: 1.2px;
}
.section-divider:first-child { margin-top: 0; border-top: none; padding-top: 0; }
@media print {
  nav.sidebar { display: none; }
  main { padding: 0; max-width: 100%; }
  section.slide { box-shadow: none; page-break-after: always; border-radius: 0; }
}
"""

JS = """
document.addEventListener('keydown', function(e) {
  if (e.target.tagName === 'INPUT' || e.target.tagName === 'TEXTAREA') return;
  const slides = Array.from(document.querySelectorAll('section.slide'));
  const cur = slides.findIndex(s => {
    const r = s.getBoundingClientRect();
    return r.top >= -50 && r.top < 200;
  });
  if (e.key === 'j' || e.key === 'ArrowDown') {
    const next = slides[Math.min(slides.length - 1, Math.max(0, cur) + 1)];
    if (next) next.scrollIntoView({behavior: 'smooth'});
    e.preventDefault();
  } else if (e.key === 'k' || e.key === 'ArrowUp') {
    const prev = slides[Math.max(0, (cur === -1 ? 0 : cur) - 1)];
    if (prev) prev.scrollIntoView({behavior: 'smooth'});
    e.preventDefault();
  }
});
"""


# ---------------------------------------------------------------------------
# 데이터 로딩 + 정리
# ---------------------------------------------------------------------------

def load_xlsx(path: Path) -> dict[str, dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, dict] = {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        rows: list[list[Any]] = []
        meta_lines: list[str] = []
        headers: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if i <= 4:
                v = row[0]
                if v:
                    meta_lines.append(str(v))
                continue
            if i == 5:
                continue  # 빈 행
            if i == 6:
                headers = [str(c) if c is not None else "" for c in row if c is not None]
                continue
            data_row = list(row[:len(headers)])
            if all(c is None for c in data_row):
                continue
            rows.append(data_row)
        sheets[sn] = {"meta": meta_lines, "headers": headers, "rows": rows}
    wb.close()
    return sheets


# ---------------------------------------------------------------------------
# HTML 렌더링
# ---------------------------------------------------------------------------

def fmt_cell(v: Any, header: str = "") -> tuple[str, str]:
    """Return (rendered_html, css_class)."""
    if v is None:
        return ("", "")
    if isinstance(v, float):
        # rate / share 컬럼은 0~1 사이 → 퍼센트
        if 0.0 <= v <= 1.0 and any(k in header.lower() for k in ("rate", "share", "비율", "%")):
            return (f"{v*100:.2f}%", "num")
        if v.is_integer():
            return (f"{int(v):,}", "num")
        return (f"{v:.4f}", "num")
    if isinstance(v, int):
        return (f"{v:,}", "num")
    s = str(v)
    if s.startswith("http"):
        short = s if len(s) <= 80 else s[:77] + "…"
        return (f'<a href="{html.escape(s)}" target="_blank" rel="noopener">{html.escape(short)}</a>', "url")
    return (html.escape(s), "")


# ---------------------------------------------------------------------------
# Wide-format pivot (PDF 모양 맞추기)
# ---------------------------------------------------------------------------

def pivot_for_pdf(sheet_id: str, headers: list[str], rows: list[list[Any]]
                  ) -> tuple[list[str], list[list[Any]]] | None:
    """PDF 와 축 방향 다른 시트 5개를 wide-format 으로 재구성. None 반환 시 변환 없음."""
    if not headers or not rows:
        return None

    def col(name): return headers.index(name) if name in headers else None

    if sheet_id == "08":
        # AI Engine 행 × Brand 컬럼. Rate 표시.
        ai_idx = col("AI Engine"); br_idx = col("Brand"); rt_idx = col("Rate")
        q_idx = col("Questions"); tbm_idx = col("Total Brand Mentions")
        if None in (ai_idx, br_idx, rt_idx): return None
        engines, brands, mat = [], [], {}
        meta = {}
        for r in rows:
            ai, br = r[ai_idx], r[br_idx]
            if ai not in engines: engines.append(ai)
            if br not in brands: brands.append(br)
            mat[(ai, br)] = r[rt_idx]
            meta[ai] = (r[q_idx] if q_idx is not None else "", r[tbm_idx] if tbm_idx is not None else "")
        new_headers = ["AI Engine", "Questions", "Total Brand Mentions"] + brands
        new_rows = []
        for ai in engines:
            row = [ai, meta[ai][0], meta[ai][1]] + [mat.get((ai, b), "") for b in brands]
            new_rows.append(row)
        return new_headers, new_rows

    if sheet_id == "12":
        # AI Platform 행 × Domain Type 컬럼. Share %.
        pl_idx = col("AI Platform"); dt_idx = col("Domain Type"); sh_idx = col("Share")
        if None in (pl_idx, dt_idx, sh_idx): return None
        platforms, dtypes, mat = [], [], {}
        for r in rows:
            pl, dt = r[pl_idx], r[dt_idx]
            if pl not in platforms: platforms.append(pl)
            if dt not in dtypes: dtypes.append(dt)
            mat[(pl, dt)] = r[sh_idx]
        new_headers = ["AI Platform"] + dtypes
        new_rows = [[pl] + [mat.get((pl, dt), "") for dt in dtypes] for pl in platforms]
        return new_headers, new_rows

    if sheet_id == "16":
        # Topic + AI Engine 행 × Brand 컬럼.
        tp_idx = col("Topic"); ai_idx = col("AI Engine"); br_idx = col("Brand"); rt_idx = col("Rate")
        q_idx = col("Questions")
        if None in (tp_idx, ai_idx, br_idx, rt_idx): return None
        keys, brands, mat, qmeta = [], [], {}, {}
        for r in rows:
            k = (r[tp_idx], r[ai_idx])
            if k not in keys: keys.append(k)
            if r[br_idx] not in brands: brands.append(r[br_idx])
            mat[(k, r[br_idx])] = r[rt_idx]
            qmeta[k] = r[q_idx] if q_idx is not None else ""
        new_headers = ["Topic", "AI Engine", "Questions"] + brands
        new_rows = [[tp, ai, qmeta[(tp, ai)]] + [mat.get(((tp, ai), b), "") for b in brands]
                    for (tp, ai) in keys]
        return new_headers, new_rows

    if sheet_id == "18":
        # Topic + Position + Axis Value 행 × Brand 컬럼. Mention Rate 표시.
        tp_idx = col("Topic"); ps_idx = col("Position (Axis)"); av_idx = col("Axis Value")
        br_idx = col("Brand"); mr_idx = col("Mention Rate"); n_idx = col("전체 응답 수")
        if None in (tp_idx, ps_idx, av_idx, br_idx, mr_idx): return None
        keys, brands, mat, nmeta = [], [], {}, {}
        for r in rows:
            k = (r[tp_idx], r[ps_idx], r[av_idx])
            if k not in keys: keys.append(k)
            if r[br_idx] not in brands: brands.append(r[br_idx])
            mat[(k, r[br_idx])] = r[mr_idx]
            nmeta[k] = r[n_idx] if n_idx is not None else ""
        new_headers = ["Topic", "Position (Axis)", "Axis Value", "전체 응답 수"] + brands
        new_rows = [[tp, ps, av, nmeta[(tp, ps, av)]] + [mat.get(((tp, ps, av), b), "") for b in brands]
                    for (tp, ps, av) in keys]
        return new_headers, new_rows

    if sheet_id == "20":
        # Topic + AI Platform 행 × Domain Type 컬럼.
        tp_idx = col("Topic"); pl_idx = col("AI Platform"); dt_idx = col("Domain Type")
        sh_idx = col("Share")
        if None in (tp_idx, pl_idx, dt_idx, sh_idx): return None
        keys, dtypes, mat = [], [], {}
        for r in rows:
            k = (r[tp_idx], r[pl_idx])
            if k not in keys: keys.append(k)
            if r[dt_idx] not in dtypes: dtypes.append(r[dt_idx])
            mat[(k, r[dt_idx])] = r[sh_idx]
        new_headers = ["Topic", "AI Platform"] + dtypes
        new_rows = [[tp, pl] + [mat.get(((tp, pl), dt), "") for dt in dtypes]
                    for (tp, pl) in keys]
        return new_headers, new_rows

    return None


def render_table(headers: list[str], rows: list[list[Any]], max_rows: int = 100) -> str:
    if not headers:
        return "<p><em>(no data)</em></p>"
    truncated = len(rows) > max_rows
    display_rows = rows[:max_rows]
    parts = ['<div class="table-wrap"><table><thead><tr>']
    for h in headers:
        parts.append(f"<th>{html.escape(str(h))}</th>")
    parts.append("</tr></thead><tbody>")
    for r in display_rows:
        # Total/Grand Total 행 강조
        first_cell = str(r[0]) if r else ""
        cls = " class=\"total\"" if "Total" in first_cell or "Grand" in first_cell else ""
        parts.append(f"<tr{cls}>")
        for j, v in enumerate(r):
            header = headers[j] if j < len(headers) else ""
            content, css = fmt_cell(v, header)
            parts.append(f'<td class="{css}">{content}</td>')
        parts.append("</tr>")
    parts.append("</tbody></table></div>")
    if truncated:
        parts.append(f'<div class="row-count">전체 {len(rows):,} 행 중 {max_rows:,} 행 표시</div>')
    else:
        parts.append(f'<div class="row-count">총 {len(rows):,} 행</div>')
    return "".join(parts)


def render_slide(spec: SheetSpec, sheet_data: dict | None) -> str:
    headers = sheet_data["headers"] if sheet_data else spec.headers
    rows = sheet_data["rows"] if sheet_data else []

    pivoted = pivot_for_pdf(spec.sheet_id, headers, rows)
    is_pivoted = pivoted is not None
    if is_pivoted:
        display_headers, display_rows = pivoted
    else:
        display_headers, display_rows = headers, rows

    parts = [f'<section class="slide" id="sheet-{spec.sheet_id}">']
    parts.append(f"<h2>{spec.sheet_id}. {html.escape(spec.title_ko)}</h2>")
    parts.append('<div class="pdf-meta">')
    parts.append(f'<span class="pill">{html.escape(spec.pdf_section)}</span>')
    parts.append(f"<span>page {html.escape(spec.pdf_pages)} · sheet <code>{html.escape(spec.sheet_name)}</code></span>")
    if is_pivoted:
        parts.append(' <span class="pill" style="background:#fff3cd;color:#856404;">PDF 축에 맞춰 wide-format 변환됨 (xlsx 는 long)</span>')
    parts.append("</div>")
    parts.append(f'<div class="purpose">{html.escape(spec.purpose)}</div>')
    parts.append(f'<details class="data-rules"><summary><strong>데이터 룰</strong></summary>{html.escape(spec.data_rules)}</details>')
    parts.append(render_table(display_headers, display_rows))
    parts.append(f'<div class="validation"><strong>검증 포인트:</strong> {html.escape(spec.validation)}</div>')
    parts.append("</section>")
    return "\n".join(parts)


def render_sidebar(specs: list[SheetSpec]) -> str:
    section_groups = [
        ("Section 1. Analysis Overview", [s for s in specs if s.sheet_id in ("01", "02", "03")]),
        ("Section 2. Overall Analysis Rate", [s for s in specs if s.sheet_id in ("04", "05", "06", "07", "08", "09", "10", "11", "12")]),
        ("Section 3. Topic Analysis", [s for s in specs if int(s.sheet_id) >= 13]),
    ]
    parts = ['<nav class="sidebar">',
             '<h1>YSL Audit Report</h1>',
             '<div class="meta">PPT 슬라이드 미리보기 · 26 시트</div>']
    for sec_title, sec_sheets in section_groups:
        parts.append(f'<a class="section-head">{html.escape(sec_title)}</a>')
        for s in sec_sheets:
            parts.append(f'<a href="#sheet-{s.sheet_id}"><span class="num">{s.sheet_id}</span>{html.escape(s.title_ko)}</a>')
    parts.append("</nav>")
    return "\n".join(parts)


def render_section_dividers(specs: list[SheetSpec], rendered_slides: dict[str, str]) -> list[str]:
    out = []
    section_breaks = {
        "01": "Section 1. Analysis Overview",
        "04": "Section 2. Overall Analysis Rate",
        "13": "Section 3. Topic Analysis",
    }
    for s in specs:
        if s.sheet_id in section_breaks:
            out.append(f'<div class="section-divider">{html.escape(section_breaks[s.sheet_id])}</div>')
        out.append(rendered_slides[s.sheet_id])
    return out


def build_html(specs: list[SheetSpec], xlsx_data: dict[str, dict], title: str) -> str:
    rendered = {s.sheet_id: render_slide(s, xlsx_data.get(s.sheet_name)) for s in specs}
    body_parts = render_section_dividers(specs, rendered)
    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{html.escape(title)}</title>
<style>{CSS}</style>
</head>
<body>
{render_sidebar(specs)}
<main>
{chr(10).join(body_parts)}
</main>
<script>{JS}</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# main
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="projects/ysl/report/Bubbleshare_YSL_Audit_Report_Data_0504.xlsx")
    parser.add_argument("--output", default=None)
    args = parser.parse_args(argv)

    inp = Path(args.input)
    if args.output:
        out = Path(args.output)
    else:
        m = re.search(r"_(\d{4})\.xlsx$", inp.name)
        suffix = f"_{m.group(1)}" if m else ""
        out = Path(f"projects/ysl/report/audit_report_mockup{suffix}.html")

    print(f"[load] {inp}")
    xlsx = load_xlsx(inp)
    print(f"[stat] {len(xlsx)} sheets loaded")

    title = f"YSL Audit Report Mockup ({inp.stem})"
    html_str = build_html(SHEETS, xlsx, title)

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html_str, encoding="utf-8")
    print(f"[write] {out} ({out.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
